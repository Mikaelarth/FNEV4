#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FNEV4 - Générateur de Template Excel pour Import Clients DGI
=============================================================

Ce script génère un template Excel conforme aux spécifications DGI FNE
pour l'import des clients avec les templates B2B, B2C, B2G, B2F.

Basé sur la documentation :
- CAHIER-DES-CHARGES-FNEV4.md
- FNE-procedureapi.md  
- ClientImportModelDgi.cs

Date: 6 Septembre 2025
Version: 1.0
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
import os

def create_client_import_template_dgi():
    """
    Crée un template Excel pour l'import des clients
    conforme aux spécifications API DGI FNE
    """
    
    # Création du workbook
    wb = Workbook()
    
    # === FEUILLE 1: TEMPLATE D'IMPORT ===
    ws_template = wb.active
    ws_template.title = "Template_Import_Clients"
    
    # En-têtes selon ClientImportModelDgi
    headers = [
        "Code_Client",           # ClientCode - Obligatoire
        "Nom_Raison_Sociale",    # Name - Obligatoire  
        "Template_Facturation",  # Template (B2B/B2C/B2G/B2F) - Obligatoire
        "NCC_Client",           # ClientNcc - Obligatoire si B2B, interdit si B2C
        "Nom_Commercial",       # CommercialName
        "Adresse",              # Address
        "Ville",                # City
        "Code_Postal",          # PostalCode
        "Pays",                 # Country
        "Telephone",            # Phone
        "Email",                # Email
        "Representant",         # Representative  
        "Numero_Fiscal",        # TaxNumber
        "Devise",               # Currency (XOF par défaut)
        "Actif",                # IsActive (Oui/Non)
        "Notes"                 # Notes
    ]
    
    # === STYLES ===
    
    # Style pour l'en-tête
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style pour les cellules obligatoires
    required_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    
    # Style pour les cellules conditionnelles
    conditional_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    # === EN-TÊTES AVEC STYLES ===
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_template.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
        
        # Largeur des colonnes
        column_letter = ws_template.cell(row=1, column=col_num).column_letter
        if header in ["Nom_Raison_Sociale", "Nom_Commercial", "Adresse"]:
            ws_template.column_dimensions[column_letter].width = 25
        elif header in ["NCC_Client", "Code_Client"]:
            ws_template.column_dimensions[column_letter].width = 15
        elif header == "Template_Facturation":
            ws_template.column_dimensions[column_letter].width = 18
        elif header == "Notes":
            ws_template.column_dimensions[column_letter].width = 30
        else:
            ws_template.column_dimensions[column_letter].width = 12
    
    # === EXEMPLES DE DONNÉES ===
    
    examples = [
        # Client B2B (Entreprise avec NCC)
        ["ENT001", "KPMG CÔTE D'IVOIRE", "B2B", "9502363N", "KPMG CI", 
         "Boulevard Angoulvant", "Abidjan", "01 BP 2651", "Côte d'Ivoire", 
         "+225 27 20 30 40 00", "info@kpmg.ci", "Directeur Général", 
         "9502363N", "XOF", "Oui", "Cabinet d'audit international"],
        
        # Client B2C (Particulier sans NCC)  
        ["PART001", "KOUAME ARTHUR LEBLANC", "B2C", "", "Arthur Le Grand",
         "Cocody Angré", "Abidjan", "08 BP 2150", "Côte d'Ivoire",
         "+225 07 08 09 10 11", "arthur.leblanc@gmail.com", "Lui-même",
         "", "XOF", "Oui", "Client particulier régulier"],
         
        # Client B2G (Gouvernemental)
        ["GOUV001", "MINISTERE DU COMMERCE", "B2G", "1000001A", "Min Commerce",
         "Plateau", "Abidjan", "BP V 65", "Côte d'Ivoire",
         "+225 27 20 25 35 00", "contact@commerce.gouv.ci", "Directeur de Cabinet",
         "1000001A", "XOF", "Oui", "Institution gouvernementale"],
         
        # Client B2F (International)  
        ["INT001", "NESTLÉ FRANCE SAS", "B2F", "FR123456789", "Nestlé France",
         "7 Boulevard Pierre Carle", "Noisiel", "77446", "France",
         "+33 1 60 53 20 00", "contact@nestle.fr", "Directeur Export Afrique",
         "FR123456789", "EUR", "Oui", "Multinationale agroalimentaire"]
    ]
    
    # Ajout des exemples
    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            cell = ws_template.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            
            # Styles conditionnels
            header_name = headers[col_num-1]
            if header_name in ["Code_Client", "Nom_Raison_Sociale", "Template_Facturation"]:
                cell.fill = required_fill
            elif header_name == "NCC_Client" and example[2] in ["B2B", "B2G", "B2F"]:
                cell.fill = conditional_fill
    
    # === VALIDATIONS DE DONNÉES ===
    
    # Validation Template (colonne C)
    template_validation = DataValidation(
        type="list",
        formula1='"B2B,B2C,B2G,B2F"',
        allow_blank=False
    )
    template_validation.error = "Template doit être B2B, B2C, B2G ou B2F"
    template_validation.errorTitle = "Template invalide"
    ws_template.add_data_validation(template_validation)
    template_validation.add(f"C2:C1000")
    
    # Validation Devise (colonne N)
    currency_validation = DataValidation(
        type="list", 
        formula1='"XOF,USD,EUR,JPY,CAD,GBP,AUD,CNH,CHF,HKD,NZD"',
        allow_blank=False
    )
    currency_validation.error = "Devise non supportée par l'API DGI"
    currency_validation.errorTitle = "Devise invalide"
    ws_template.add_data_validation(currency_validation)
    currency_validation.add(f"N2:N1000")
    
    # Validation Actif (colonne O)
    active_validation = DataValidation(
        type="list",
        formula1='"Oui,Non"',
        allow_blank=False
    )
    ws_template.add_data_validation(active_validation)
    active_validation.add(f"O2:O1000")
    
    # === COMMENTAIRES EXPLICATIFS ===
    
    # Commentaire sur Template
    template_comment = Comment(
        "TEMPLATES API DGI FNE:\n\n" +
        "• B2B: Entreprise avec NCC (NCC obligatoire)\n" +
        "• B2C: Particulier (NCC interdit)\n" + 
        "• B2G: Gouvernemental (NCC obligatoire)\n" +
        "• B2F: International (devise étrangère obligatoire)",
        "FNEV4"
    )
    ws_template["C1"].comment = template_comment
    
    # Commentaire sur NCC
    ncc_comment = Comment(
        "NCC CLIENT:\n\n" +
        "• Obligatoire pour B2B, B2G, B2F\n" +
        "• Interdit pour B2C\n" +
        "• Format: 8-11 caractères alphanumériques\n" +
        "• Exemple: 9502363N",
        "FNEV4"
    )
    ws_template["D1"].comment = ncc_comment
    
    # === FEUILLE 2: DOCUMENTATION ===
    ws_doc = wb.create_sheet("Documentation")
    
    documentation = [
        ["DOCUMENTATION TEMPLATE IMPORT CLIENTS DGI"],
        [""],
        ["1. OBJECTIF"],
        ["Ce template permet d'importer des clients dans FNEV4"],
        ["conformément aux spécifications API DGI FNE 2025."],
        [""],
        ["2. TEMPLATES DE FACTURATION"],
        ["B2B: Business to Business (Entreprise → Entreprise)"],
        ["  - Client = Entreprise avec NCC"],
        ["  - NCC obligatoire"],
        ["  - Devise = XOF généralement"],
        [""],
        ["B2C: Business to Consumer (Entreprise → Particulier)"],
        ["  - Client = Particulier sans NCC"],
        ["  - NCC interdit"],
        ["  - Devise = XOF uniquement"],
        [""],
        ["B2G: Business to Government (Entreprise → Gouvernement)"],
        ["  - Client = Institution gouvernementale"],
        ["  - NCC obligatoire"],
        ["  - Devise = XOF uniquement"],
        [""],
        ["B2F: Business to Foreign (Entreprise → Étranger)"],
        ["  - Client = Entreprise étrangère"],
        ["  - NCC obligatoire (identification internationale)"],
        ["  - Devise étrangère obligatoire (≠ XOF)"],
        ["  - Pays ≠ Côte d'Ivoire"],
        [""],
        ["3. RÈGLES DE VALIDATION"],
        [""],
        ["CHAMPS OBLIGATOIRES:"],
        ["  - Code_Client (unique)"],
        ["  - Nom_Raison_Sociale"],
        ["  - Template_Facturation"],
        [""],
        ["RÈGLES NCC:"],
        ["  - Format: 8 à 11 caractères alphanumériques"],
        ["  - Obligatoire: B2B, B2G, B2F"],
        ["  - Interdit: B2C"],
        [""],
        ["DEVISES SUPPORTÉES:"],
        ["  - XOF: Franc CFA (par défaut)"],
        ["  - USD: Dollar Américain"],
        ["  - EUR: Euro"],
        ["  - JPY: Yen Japonais"],
        ["  - CAD: Dollar Canadien"],
        ["  - GBP: Livre Sterling"],
        ["  - AUD: Dollar Australien"],
        ["  - CNH: Yuan Chinois"],
        ["  - CHF: Franc Suisse"],
        ["  - HKD: Dollar Hong Kong"],
        ["  - NZD: Dollar Néo-Zélandais"],
        [""],
        ["4. EXEMPLES FOURNIS"],
        ["Ligne 2: Client B2B (KPMG - Entreprise ivoirienne)"],
        ["Ligne 3: Client B2C (Arthur - Particulier)"],
        ["Ligne 4: Client B2G (Ministère - Gouvernement)"],
        ["Ligne 5: Client B2F (Nestlé - International)"],
        [""],
        ["5. IMPORT DANS FNEV4"],
        ["1. Remplir le template avec vos données"],
        ["2. Menu: Gestion Clients → Liste des clients"],
        ["3. Bouton: Import Excel"],
        ["4. Sélectionner ce fichier"],
        ["5. Configurer les options d'import"],
        ["6. Lancer l'import"],
        [""],
        ["6. SUPPORT"],
        ["En cas de problème: support.fne@dgi.gouv.ci"],
        ["Documentation complète: FNE-procedureapi.md"]
    ]
    
    for row_num, line in enumerate(documentation, 1):
        cell = ws_doc.cell(row=row_num, column=1, value=line[0])
        if row_num == 1:  # Titre
            cell.font = Font(bold=True, size=14, color="2E7D32")
        elif any(keyword in line[0] for keyword in ["1.", "2.", "3.", "4.", "5.", "6."]):  # Sections
            cell.font = Font(bold=True, size=12, color="1976D2")
        elif line[0].endswith(":"):  # Sous-sections
            cell.font = Font(bold=True, size=10)
    
    # Largeur colonne documentation
    ws_doc.column_dimensions['A'].width = 80
    
    # === FEUILLE 3: CODES ERREUR ===
    ws_errors = wb.create_sheet("Codes_Erreur")
    
    error_codes = [
        ["Code", "Type", "Description", "Solution"],
        ["VAL001", "Template", "Template invalide", "Utiliser B2B, B2C, B2G ou B2F"],
        ["VAL002", "NCC", "NCC obligatoire pour B2B", "Renseigner le NCC du client entreprise"],
        ["VAL003", "NCC", "NCC interdit pour B2C", "Laisser vide pour les particuliers"],
        ["VAL004", "NCC", "Format NCC invalide", "8-11 caractères alphanumériques"],
        ["VAL005", "Devise", "Devise non supportée", "Utiliser une devise de la liste"],
        ["VAL006", "B2F", "Devise XOF interdite pour B2F", "Utiliser une devise étrangère"],
        ["VAL007", "B2F", "Pays invalide pour B2F", "Client doit être à l'étranger"],
        ["VAL008", "Email", "Format email invalide", "Vérifier la syntaxe email"],
        ["VAL009", "Doublon", "Code client existe déjà", "Utiliser un code unique"],
        ["VAL010", "Champ", "Champ obligatoire manquant", "Remplir tous les champs requis"]
    ]
    
    for row_num, row_data in enumerate(error_codes, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_errors.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:  # En-tête
                cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            cell.border = border
    
    # Largeurs colonnes erreurs
    ws_errors.column_dimensions['A'].width = 8   # Code
    ws_errors.column_dimensions['B'].width = 12  # Type  
    ws_errors.column_dimensions['C'].width = 35  # Description
    ws_errors.column_dimensions['D'].width = 40  # Solution
    
    # === SAUVEGARDE ===
    
    filename = "modele_import_clients_dgi.xlsx"
    filepath = os.path.join("data", "templates", filename)
    
    # Créer le dossier si nécessaire
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    # Sauvegarder
    wb.save(filepath)
    
    print(f"✅ Template Excel créé: {filepath}")
    print(f"📊 3 feuilles: Template, Documentation, Codes d'erreur")
    print(f"📝 4 exemples clients (B2B, B2C, B2G, B2F)")
    print(f"✅ Validations de données intégrées")
    print(f"📚 Documentation complète incluse")
    
    return filepath

def create_test_client_data():
    """
    Crée un fichier avec plus d'exemples de clients pour test
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test_Clients_DGI"
    
    # En-têtes
    headers = [
        "Code_Client", "Nom_Raison_Sociale", "Template_Facturation", "NCC_Client",
        "Nom_Commercial", "Adresse", "Ville", "Code_Postal", "Pays",
        "Telephone", "Email", "Representant", "Numero_Fiscal", 
        "Devise", "Actif", "Notes"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Données de test étendues
    test_data = [
        # Clients B2B (Entreprises ivoiriennes)
        ["B2B001", "ORANGE CÔTE D'IVOIRE", "B2B", "9501234A", "Orange CI", 
         "Boulevard Lagunaire", "Abidjan", "01 BP 2020", "Côte d'Ivoire",
         "+225 27 20 20 20 20", "contact@orange.ci", "DG Orange", 
         "9501234A", "XOF", "Oui", "Opérateur télécom"],
         
        ["B2B002", "BANQUE ATLANTIQUE", "B2B", "9502345B", "BACI", 
         "Boulevard de la République", "Abidjan", "01 BP 1275", "Côte d'Ivoire",
         "+225 27 20 30 40 50", "info@baci.ci", "Directeur Général", 
         "9502345B", "XOF", "Oui", "Banque commerciale"],
         
        ["B2B003", "COMPAGNIE IVOIRIENNE D'ELECTRICITE", "B2B", "9503456C", "CIE", 
         "Avenue Christiani", "Abidjan", "01 BP 6923", "Côte d'Ivoire",
         "+225 27 21 23 45 67", "service.client@cie.ci", "Directeur Commercial", 
         "9503456C", "XOF", "Oui", "Distribution électricité"],
        
        # Clients B2C (Particuliers)
        ["PART001", "TRAORE MOUSSA", "B2C", "", "Moussa Traoré",
         "Cocody Riviera", "Abidjan", "08 BP 1500", "Côte d'Ivoire",
         "+225 07 01 02 03 04", "moussa.traore@yahoo.fr", "Lui-même",
         "", "XOF", "Oui", "Commerçant particulier"],
         
        ["PART002", "KONE AMINATA", "B2C", "", "Aminata Koné",
         "Yopougon Selmer", "Abidjan", "23 BP 500", "Côte d'Ivoire",
         "+225 05 06 07 08 09", "aminata.kone@gmail.com", "Elle-même",
         "", "XOF", "Oui", "Enseignante"],
         
        ["PART003", "OUATTARA IBRAHIM", "B2C", "", "Ibrahim Ouattara",
         "Marcory Zone 4", "Abidjan", "18 BP 2500", "Côte d'Ivoire",
         "+225 01 23 45 67 89", "ibrahim.ouattara@hotmail.com", "Lui-même",
         "", "XOF", "Oui", "Ingénieur informatique"],
        
        # Clients B2G (Gouvernementaux)  
        ["GOUV001", "MINISTERE DE L'EDUCATION NATIONALE", "B2G", "1000001A", "MEN",
         "Plateau Gouvernement", "Abidjan", "BP V 120", "Côte d'Ivoire",
         "+225 27 20 21 40 00", "contact@men.gouv.ci", "Secrétaire Général",
         "1000001A", "XOF", "Oui", "Ministère Education"],
         
        ["GOUV002", "CONSEIL GENERAL DE YAMOUSSOUKRO", "B2G", "1000002B", "CG Yamoussoukro",
         "Avenue Général de Gaulle", "Yamoussoukro", "BP 825", "Côte d'Ivoire",
         "+225 30 64 05 06", "contact@cg-yamoussoukro.ci", "Président",
         "1000002B", "XOF", "Oui", "Collectivité territoriale"],
        
        # Clients B2F (Internationaux)
        ["INT001", "TOTAL ENERGIES FRANCE", "B2F", "FR789012345", "Total Energies",
         "Tour Total", "Paris La Défense", "92800", "France",
         "+33 1 47 44 45 46", "contact@totalenergies.fr", "Directeur Afrique",
         "FR789012345", "EUR", "Oui", "Pétrolier international"],
         
        ["INT002", "UNILEVER DEUTSCHLAND GMBH", "B2F", "DE987654321", "Unilever DE",
         "Strandkai 1", "Hamburg", "20457", "Allemagne",
         "+49 40 3493 0", "info.de@unilever.com", "Export Manager",
         "DE987654321", "EUR", "Oui", "Biens de consommation"],
         
        ["INT003", "WALMART INC", "B2F", "US123456789", "Walmart",
         "702 SW 8th Street", "Bentonville", "72716", "États-Unis",
         "+1 479 273 4000", "international@walmart.com", "Global Sourcing",
         "US123456789", "USD", "Oui", "Grande distribution"],
         
        ["INT004", "MITSUBISHI CORPORATION", "B2F", "JP456789123", "Mitsubishi Corp",
         "Marunouchi 2-3-1", "Tokyo", "100-8086", "Japon",
         "+81 3 3210 2121", "info@mitsubishi.com", "Africa Division",
         "JP456789123", "JPY", "Oui", "Trading international"]
    ]
    
    # Ajouter les données
    for row_num, row_data in enumerate(test_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Sauvegarder
    filename = "test_clients_dgi_complet.xlsx"
    filepath = os.path.join("data", "templates", filename)
    wb.save(filepath)
    
    print(f"✅ Fichier de test créé: {filepath}")
    print(f"📊 {len(test_data)} clients test (B2B: 3, B2C: 3, B2G: 2, B2F: 4)")
    
    return filepath

if __name__ == "__main__":
    print("🚀 FNEV4 - Générateur Template Import Clients DGI")
    print("=" * 60)
    
    # Créer le template principal
    template_path = create_client_import_template_dgi()
    
    print()
    
    # Créer le fichier de test
    test_path = create_test_client_data()
    
    print()
    print("📋 Fichiers créés:")
    print(f"  1. Template: {template_path}")
    print(f"  2. Test data: {test_path}")
    print()
    print("🔧 Utilisation:")
    print("  1. Ouvrir le template dans Excel")
    print("  2. Remplacer les exemples par vos données")
    print("  3. Respecter les règles de validation")
    print("  4. Importer dans FNEV4")
    print()
    print("📚 Documentation complète dans l'onglet 'Documentation'")
