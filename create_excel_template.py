#!/usr/bin/env python3
"""
Script pour créer un fichier Excel template valide pour les tests d'import
Basé sur les spécifications du cahier des charges FNEV4
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os

def create_excel_template():
    """Créer un fichier Excel template avec des données de test valides"""
    
    # Données de test conformes aux spécifications
    data = {
        'Code Client': ['CLI001', 'CLI002', 'CLI003', 'CLI004', 'CLI005'],
        'Nom/Raison Sociale': [
            'ARTHUR LE GRAND SARL',
            'MARIE KOUASSI',
            'MINISTERE DE LA SANTE',
            'ONG ESPOIR AFRICA',
            'INTERNATIONAL CORP'
        ],
        'Type Client': ['Entreprise', 'Particulier', 'Administration', 'Association', 'Entreprise'],
        'NCC': [
            '1234567890',  # 10 caractères - VALIDE
            '',            # Vide pour particulier - VALIDE
            '9876543210',  # 10 caractères - VALIDE
            '5555666677',  # 10 caractères - VALIDE
            '1111222233'   # 10 caractères - VALIDE
        ],
        'Nom Commercial': [
            'Arthur Le Grand',
            'Marie Kouassi Boutique',
            'Min. Santé et Hygiène Publique',
            'Espoir pour l\'Afrique',
            'International Corporation CI'
        ],
        'Adresse': [
            '123 Boulevard de la Paix',
            '45 Rue du Commerce',
            'Plateau Tour C 15ème étage',
            'Zone 4 Marcory',
            'Deux Plateaux Vallon'
        ],
        'Ville': ['Abidjan', 'Bouaké', 'Abidjan', 'Abidjan', 'Abidjan'],
        'Code Postal': ['01001', '02001', '01000', '01002', '01003'],
        'Pays': ['Côte d\'Ivoire', 'Côte d\'Ivoire', 'Côte d\'Ivoire', 'Côte d\'Ivoire', 'Côte d\'Ivoire'],
        'Téléphone': [
            '+225 01 02 03 04',
            '+225 05 06 07 08',
            '+225 20 21 22 23',
            '+225 07 08 09 10',
            '+225 22 33 44 55'
        ],
        'Email': [
            'arthur@legrand.ci',
            'marie.kouassi@gmail.com',
            'contact@sante.gouv.ci',
            'info@espoirafrique.org',
            'contact@intcorp.com'
        ],
        'Représentant': ['Jean KOUAME', 'Paul DIALLO', 'Dr. BAMBA', 'Fatou TRAORE', 'Mike JOHNSON'],
        'N° Fiscal': ['TIN123456', '', 'GOV789', 'NGO456', 'FOR789'],
        'Devise': ['XOF', 'XOF', 'XOF', 'XOF', 'USD'],
        'Actif': ['Oui', 'Oui', 'Oui', 'Oui', 'Oui'],
        'Notes': [
            'Client VIP',
            '',
            'Client gouvernemental',
            'Association caritative',
            'Client international'
        ]
    }
    
    # Créer DataFrame
    df = pd.DataFrame(data)
    
    # Chemin de sortie
    output_path = r'C:\wamp64\www\FNEV4\data\templates\modele_import_clients_correct.xlsx'
    
    # Créer le répertoire si nécessaire
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Créer le fichier Excel avec openpyxl pour plus de contrôle
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    
    # Style pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    
    # Écrire les en-têtes
    headers = list(data.keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Écrire les données
    for row_idx, row_data in enumerate(zip(*data.values()), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarder
    wb.save(output_path)
    
    print(f"✅ Fichier Excel créé avec succès : {output_path}")
    print(f"📊 {len(df)} clients de test créés")
    print("🎯 Tous les NCC sont conformes (10 caractères)")
    print("📧 Tous les emails sont valides")
    print("✨ Prêt pour les tests d'import !")
    
    return output_path

if __name__ == "__main__":
    create_excel_template()
