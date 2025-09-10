#!/usr/bin/env python3
"""
Script pour créer un fichier Excel de test pour FNEV4 - Import Sage 100
Version avec pourcentages TVA comme dans les vrais fichiers Sage 100
"""

import openpyxl
from openpyxl import Workbook

def creer_fichier_test():
    """Crée un fichier Excel de test avec des factures Sage 100 utilisant des pourcentages TVA"""
    
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # === FACTURE 1: Facture normale avec client valide ===
    ws1 = wb.create_sheet("Facture_001")
    
    # Entête de la facture
    ws1['A3'] = "FAC123"  # Numéro de facture
    ws1['A5'] = "CLI001"  # Code client
    ws1['A6'] = "1234567A"  # NCC client (valide)
    ws1['A8'] = "45509"  # Date (numéro de série Excel pour 2024-09-08)
    ws1['A10'] = "Abidjan"  # Point de vente
    ws1['A11'] = "SOCIÉTÉ ABC SARL"  # Nom du client
    ws1['A18'] = "cash"  # Moyen de paiement
    
    # Données des produits avec pourcentages TVA
    produits_data = [
        ("PROD001", "Produit Standard TVA", 100000, 2, "pcs", 18, 200000),    # 18%
        ("PROD002", "Service TVAB", 150000, 1, "service", 9, 150000),        # 9%
        ("PROD003", "Produit Exonéré TVAC", 80000, 3, "pcs", 0, 240000),     # 0%
        ("PROD004", "Export TVAD", 200000, 1, "kg", 0, 200000),              # 0%
        ("PROD005", "Autre Produit TVA", 120000, 2, "pcs", 18, 240000),      # 18%
    ]
    
    # Remplir les données des produits
    for i, (code, designation, prix, quantite, unite, tva_pct, montant) in enumerate(produits_data, 20):
        ws1[f'B{i}'] = code
        ws1[f'C{i}'] = designation
        ws1[f'D{i}'] = prix
        ws1[f'E{i}'] = quantite
        ws1[f'F{i}'] = unite
        ws1[f'G{i}'] = tva_pct  # Pourcentage TVA (18, 9, 0)
        ws1[f'H{i}'] = montant
    
    # === FACTURE 2: Client divers (code 1999) ===
    ws2 = wb.create_sheet("Facture_002")
    
    # Entête
    ws2['A3'] = "FAC124"
    ws2['A5'] = "1999"  # Code client divers
    ws2['A6'] = ""  # Pas de NCC pour client divers
    ws2['A8'] = "45510"  # Date suivante
    ws2['A10'] = "San Pedro"
    ws2['A11'] = "DIVERS CLIENTS"
    ws2['A13'] = "M. Jean KOUASSI"  # Nom réel du client divers
    ws2['A18'] = "bank-transfer"
    
    # Produit simple
    ws2['B20'] = "SERV001"
    ws2['C20'] = "Prestation de service"
    ws2['D20'] = 250000
    ws2['E20'] = 1
    ws2['F20'] = "service"
    ws2['G20'] = 9  # 9% TVAB
    ws2['H20'] = 250000
    
    # === FACTURE 3: Facture avec erreurs (pour tests de validation) ===
    ws3 = wb.create_sheet("Facture_ERREURS")
    
    # Entête avec erreurs
    ws3['A3'] = ""  # Numéro de facture manquant
    ws3['A5'] = "CLIINEX"  # Code client inexistant
    ws3['A6'] = "123"  # NCC invalide (trop court)
    ws3['A8'] = "date_invalide"  # Date non numérique
    ws3['A10'] = "Bouaké"
    ws3['A11'] = "DIVERS CLIENTS"
    ws3['A15'] = "123ABC"  # NCC invalide (mauvais format)
    ws3['A18'] = "bitcoin"  # Moyen de paiement invalide
    
    # Produits avec erreurs
    ws3['B20'] = "PROD001"
    ws3['C20'] = ""  # Désignation manquante
    ws3['D20'] = "prix_invalide"  # Prix non numérique
    ws3['E20'] = 0  # Quantité nulle
    ws3['F20'] = "pcs"
    ws3['G20'] = 25  # Pourcentage TVA invalide (pas dans 0, 9, 18)
    ws3['H20'] = 100000
    
    # === FACTURE 4: Facture complexe avec tous les types de TVA ===
    ws4 = wb.create_sheet("Facture_COMPLETE")
    
    # Entête
    ws4['A3'] = "FAC999"
    ws4['A5'] = "ENTREP001"
    ws4['A6'] = "9876543Z"
    ws4['A8'] = "45511"
    ws4['A10'] = "Yamoussoukro"
    ws4['A11'] = "GRANDE ENTREPRISE SARL"
    ws4['A18'] = "check"
    
    # Produits avec tous les types de TVA
    produits_complets = [
        ("SERV001", "Service standard", 100000, 1, "service", 18, 100000),      # TVA 18%
        ("SERV002", "Service réduit", 200000, 1, "service", 9, 200000),         # TVAB 9%
        ("PROD001", "Produit exonéré", 150000, 2, "pcs", 0, 300000),            # TVAC 0%
        ("EXP001", "Produit export", 300000, 1, "kg", 0, 300000),               # TVAD 0%
        ("MAT001", "Matériel bureau", 80000, 5, "pcs", 18, 400000),             # TVA 18%
        ("FORM001", "Formation", 120000, 8, "heures", 9, 960000),               # TVAB 9%
    ]
    
    for i, (code, design, prix, qte, unite, tva_pct, montant) in enumerate(produits_complets, 20):
        ws4[f'B{i}'] = code
        ws4[f'C{i}'] = design
        ws4[f'D{i}'] = prix
        ws4[f'E{i}'] = qte
        ws4[f'F{i}'] = unite
        ws4[f'G{i}'] = tva_pct
        ws4[f'H{i}'] = montant
    
    # Sauvegarder le fichier
    nom_fichier = "test_factures_sage100.xlsx"
    wb.save(nom_fichier)
    
    print(f"✅ Fichier de test créé: {nom_fichier}")
    print()
    print("📊 Contenu du fichier (avec pourcentages TVA):")
    print("   • Facture_001: Client normal avec TVA 18%, 9%, 0%")
    print("   • Facture_002: Client divers avec TVA 9%")
    print("   • Facture_ERREURS: Tests de validation")
    print("   • Facture_COMPLETE: Tous types de TVA")
    print()
    print("🔢 Correspondance TVA:")
    print("   • 18% → TVA (Taxe sur la Valeur Ajoutée)")
    print("   • 9%  → TVAB (TVA au taux réduit)")
    print("   • 0%  → TVAC (TVA exonérée) ou TVAD (Export/TEE/RME)")
    print()
    print("📍 Fichier sauvegardé dans:", nom_fichier)

if __name__ == "__main__":
    creer_fichier_test()
