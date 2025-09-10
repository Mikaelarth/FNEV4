#!/usr/bin/env python3
"""
Créateur de fichier de test pour validation des règles NCC mises à jour
B2B: NCC obligatoire
Divers: NCC optionnel (template B2C par défaut)
"""

import openpyxl
from datetime import datetime

def create_updated_ncc_test():
    """Crée un fichier de test avec les nouvelles règles NCC"""
    
    workbook = openpyxl.Workbook()
    
    # Supprimer la feuille par défaut
    workbook.remove(workbook.active)
    
    # Test 1: Client B2B avec NCC (OK)
    sheet1 = workbook.create_sheet("B2B_Avec_NCC")
    
    sheet1['A3'] = "556001"  # Numéro facture
    sheet1['A5'] = "2000"    # Code client B2B
    sheet1['A6'] = "1234567A"  # NCC client B2B (obligatoire)
    sheet1['A8'] = "2025-09-10"
    sheet1['A10'] = "Gestoci"
    sheet1['A11'] = "ENTREPRISE ABC SARL"
    sheet1['A18'] = "VIREMENT"
    
    sheet1['B20'] = "PROD001"
    sheet1['C20'] = "Service consulting"
    sheet1['D20'] = 1000000
    sheet1['E20'] = 1
    sheet1['F20'] = "service"
    sheet1['G20'] = "TVA"
    sheet1['H20'] = 1000000
    
    # Test 2: Client B2B SANS NCC (ERREUR)
    sheet2 = workbook.create_sheet("B2B_Sans_NCC")
    
    sheet2['A3'] = "556002"
    sheet2['A5'] = "3000"    # Code client B2B
    # A6 vide - ERREUR car obligatoire pour B2B
    sheet2['A8'] = "2025-09-10"
    sheet2['A10'] = "Gestoci"
    sheet2['A11'] = "ENTREPRISE XYZ SARL"
    sheet2['A18'] = "VIREMENT"
    
    sheet2['B20'] = "PROD002"
    sheet2['C20'] = "Formation"
    sheet2['D20'] = 500000
    sheet2['E20'] = 2
    sheet2['F20'] = "session"
    sheet2['G20'] = "TVA"
    sheet2['H20'] = 1000000
    
    # Test 3: Client Divers avec NCC (OK)
    sheet3 = workbook.create_sheet("Divers_Avec_NCC")
    
    sheet3['A3'] = "556003"
    sheet3['A5'] = "1999"    # Code client divers
    sheet3['A6'] = "2354552Q"  # NCC générique optionnel
    sheet3['A8'] = "2025-09-10"
    sheet3['A10'] = "Gestoci"
    sheet3['A11'] = "DIVERS CLIENTS"
    sheet3['A13'] = "MARIE KOUAME"  # Nom réel obligatoire
    sheet3['A15'] = "7890123D"  # NCC spécifique optionnel
    sheet3['A18'] = "ESPECES"
    
    sheet3['B20'] = "PROD003"
    sheet3['C20'] = "Carburant"
    sheet3['D20'] = 750
    sheet3['E20'] = 50
    sheet3['F20'] = "litres"
    sheet3['G20'] = "TVA"
    sheet3['H20'] = 37500
    
    # Test 4: Client Divers SANS NCC (OK - template B2C par défaut)
    sheet4 = workbook.create_sheet("Divers_Sans_NCC")
    
    sheet4['A3'] = "556004"
    sheet4['A5'] = "1999"    # Code client divers
    # A6 et A15 vides - OK car template B2C par défaut
    sheet4['A8'] = "2025-09-10"
    sheet4['A10'] = "Gestoci"
    sheet4['A11'] = "DIVERS CLIENTS"
    sheet4['A13'] = "JEAN KOUASSI"  # Nom réel obligatoire
    sheet4['A18'] = "ESPECES"
    
    sheet4['B20'] = "PROD004"
    sheet4['C20'] = "Essence"
    sheet4['D20'] = 780
    sheet4['E20'] = 40
    sheet4['F20'] = "litres"
    sheet4['G20'] = "TVA"
    sheet4['H20'] = 31200
    
    # Test 5: Client Divers SANS nom réel (ERREUR)
    sheet5 = workbook.create_sheet("Divers_Sans_Nom")
    
    sheet5['A3'] = "556005"
    sheet5['A5'] = "1999"    # Code client divers
    sheet5['A8'] = "2025-09-10"
    sheet5['A10'] = "Gestoci"
    sheet5['A11'] = "DIVERS CLIENTS"
    # A13 vide - ERREUR car nom réel obligatoire pour divers
    sheet5['A18'] = "ESPECES"
    
    sheet5['B20'] = "PROD005"
    sheet5['C20'] = "Gazoil"
    sheet5['D20'] = 800
    sheet5['E20'] = 30
    sheet5['F20'] = "litres"
    sheet5['G20'] = "TVA"
    sheet5['H20'] = 24000
    
    # Test 6: Client Divers avec NCC invalide (ERREUR)
    sheet6 = workbook.create_sheet("Divers_NCC_Invalide")
    
    sheet6['A3'] = "556006"
    sheet6['A5'] = "1999"    # Code client divers
    sheet6['A8'] = "2025-09-10"
    sheet6['A10'] = "Gestoci"
    sheet6['A11'] = "DIVERS CLIENTS"
    sheet6['A13'] = "PAUL DIABATE"
    sheet6['A15'] = "ABC123"  # Format NCC invalide
    sheet6['A18'] = "ESPECES"
    
    sheet6['B20'] = "PROD006"
    sheet6['C20'] = "Lubrifiant"
    sheet6['D20'] = 5000
    sheet6['E20'] = 5
    sheet6['F20'] = "litres"
    sheet6['G20'] = "TVA"
    sheet6['H20'] = 25000
    
    # Sauvegarder le fichier
    filename = "test_ncc_rules_updated.xlsx"
    workbook.save(filename)
    
    print(f"✅ Fichier de test créé: {filename}")
    print("\n📋 Règles testées:")
    print("  ✅ B2B_Avec_NCC: Client B2B avec NCC (valide)")
    print("  ❌ B2B_Sans_NCC: Client B2B sans NCC (erreur)")
    print("  ✅ Divers_Avec_NCC: Client divers avec NCC (valide)")
    print("  ✅ Divers_Sans_NCC: Client divers sans NCC (valide - B2C par défaut)")
    print("  ❌ Divers_Sans_Nom: Client divers sans nom réel (erreur)")
    print("  ❌ Divers_NCC_Invalide: Client divers avec NCC format invalide (erreur)")
    
    return filename

if __name__ == "__main__":
    create_updated_ncc_test()
