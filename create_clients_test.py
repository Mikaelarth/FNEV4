#!/usr/bin/env python3
"""
Créateur de fichier de test avec validation de clients dans la base de données
"""

import openpyxl
import sqlite3
import os
from datetime import datetime

def create_clients_test_db():
    """Crée une base de données de test avec des clients"""
    db_path = "test_clients.db"
    
    # Supprimer la DB existante si elle existe
    if os.path.exists(db_path):
        os.remove(db_path)
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Créer la table Clients
    cursor.execute("""
        CREATE TABLE Clients (
            Id INTEGER PRIMARY KEY AUTOINCREMENT,
            CodeClient TEXT UNIQUE NOT NULL,
            NomCommercial TEXT NOT NULL,
            NCC TEXT,
            Template TEXT NOT NULL,
            EstActif INTEGER NOT NULL DEFAULT 1
        )
    """)
    
    # Insérer des clients de test
    clients_test = [
        ('2000', 'ENTREPRISE ABC SARL', '1234567A', 'B2B', 1),
        ('3000', 'SOCIETE XYZ EURL', '2345678B', 'B2B', 1),
        ('4000', 'COMPAGNIE DEF SA', '3456789C', 'B2B', 1),
        ('5000', 'BUSINESS GHI SARL', '4567890D', 'B2C', 1),
        ('6000', 'INACTIVE CLIENT', '5678901E', 'B2B', 0),  # Client inactif
        ('3001A', 'VENTE AU COMPTANT', '6789012F', 'B2C', 1),
        ('3001B', 'CLIENTS COMPTANT', '7890123G', 'B2C', 1),
    ]
    
    cursor.executemany("""
        INSERT INTO Clients (CodeClient, NomCommercial, NCC, Template, EstActif)
        VALUES (?, ?, ?, ?, ?)
    """, clients_test)
    
    conn.commit()
    conn.close()
    
    print(f"✅ Base de données de test créée: {db_path}")
    return db_path

def create_clients_validation_test():
    """Crée un fichier Excel de test pour la validation des clients"""
    
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    
    # Test 1: Client B2B existant et actif
    sheet1 = workbook.create_sheet("Facture_Client_B2B_OK")
    sheet1['A3'] = "556001"
    sheet1['A5'] = "2000"  # Client existant
    sheet1['A6'] = "1234567A"  # NCC correct
    sheet1['A8'] = "2025-09-10"
    sheet1['A10'] = "Gestoci"
    sheet1['A11'] = "ENTREPRISE ABC SARL"  # Nom ligne 11 pour B2B
    sheet1['A18'] = "VIREMENT"
    # Produits
    sheet1['B20'] = "PROD001"
    sheet1['C20'] = "Service consulting"
    sheet1['D20'] = 1000000
    sheet1['E20'] = 1
    sheet1['F20'] = "service"
    sheet1['G20'] = "TVA"
    sheet1['H20'] = 1000000
    
    # Test 2: Client inexistant (erreur)
    sheet2 = workbook.create_sheet("Facture_Client_Inexistant")
    sheet2['A3'] = "556002"
    sheet2['A5'] = "9999"  # Client qui n'existe pas
    sheet2['A6'] = "1234567A"
    sheet2['A8'] = "2025-09-10"
    sheet2['A10'] = "Gestoci"
    sheet2['A11'] = "CLIENT INEXISTANT"
    sheet2['A18'] = "VIREMENT"
    # Produits
    sheet2['B20'] = "PROD002"
    sheet2['C20'] = "Service inexistant"
    sheet2['D20'] = 500000
    sheet2['E20'] = 1
    sheet2['F20'] = "service"
    sheet2['G20'] = "TVA"
    sheet2['H20'] = 500000
    
    # Test 3: Client inactif (erreur)
    sheet3 = workbook.create_sheet("Facture_Client_Inactif")
    sheet3['A3'] = "556003"
    sheet3['A5'] = "6000"  # Client inactif
    sheet3['A6'] = "5678901E"
    sheet3['A8'] = "2025-09-10"
    sheet3['A10'] = "Gestoci"
    sheet3['A11'] = "INACTIVE CLIENT"
    sheet3['A18'] = "VIREMENT"
    # Produits
    sheet3['B20'] = "PROD003"
    sheet3['C20'] = "Service inactif"
    sheet3['D20'] = 300000
    sheet3['E20'] = 1
    sheet3['F20'] = "service"
    sheet3['G20'] = "TVA"
    sheet3['H20'] = 300000
    
    # Test 4: Client divers avec NCC (OK)
    sheet4 = workbook.create_sheet("Facture_Divers_Avec_NCC")
    sheet4['A3'] = "556004"
    sheet4['A5'] = "1999"  # Client divers
    sheet4['A6'] = "2354552Q"  # NCC générique
    sheet4['A8'] = "2025-09-10"
    sheet4['A10'] = "Gestoci"
    sheet4['A11'] = "DIVERS CLIENTS"
    sheet4['A13'] = "MARIE KOUAME"  # Nom réel ligne 13
    sheet4['A15'] = "7890123D"  # NCC spécifique optionnel
    sheet4['A18'] = "ESPECES"
    # Produits
    sheet4['B20'] = "PROD004"
    sheet4['C20'] = "Carburant"
    sheet4['D20'] = 750
    sheet4['E20'] = 50
    sheet4['F20'] = "litres"
    sheet4['G20'] = "TVA"
    sheet4['H20'] = 37500
    
    # Test 5: Client divers sans NCC (OK - template B2C par défaut)
    sheet5 = workbook.create_sheet("Facture_Divers_Sans_NCC")
    sheet5['A3'] = "556005"
    sheet5['A5'] = "1999"  # Client divers
    sheet5['A6'] = "2354552Q"  # NCC générique
    sheet5['A8'] = "2025-09-10"
    sheet5['A10'] = "Gestoci"
    sheet5['A11'] = "DIVERS CLIENTS"
    sheet5['A13'] = "JEAN KOUASSI"  # Nom réel ligne 13
    # A15 vide - pas de NCC spécifique (OK pour divers)
    sheet5['A18'] = "ESPECES"
    # Produits
    sheet5['B20'] = "PROD005"
    sheet5['C20'] = "Essence"
    sheet5['D20'] = 780
    sheet5['E20'] = 40
    sheet5['F20'] = "litres"
    sheet5['G20'] = "TVA"
    sheet5['H20'] = 31200
    
    # Test 6: Client B2B template B2C
    sheet6 = workbook.create_sheet("Facture_B2B_Template_B2C")
    sheet6['A3'] = "556006"
    sheet6['A5'] = "5000"  # Client avec template B2C
    sheet6['A6'] = "4567890D"
    sheet6['A8'] = "2025-09-10"
    sheet6['A10'] = "Gestoci"
    sheet6['A11'] = "BUSINESS GHI SARL"
    sheet6['A18'] = "CARTE"
    # Produits
    sheet6['B20'] = "PROD006"
    sheet6['C20'] = "Produit B2C"
    sheet6['D20'] = 150000
    sheet6['E20'] = 2
    sheet6['F20'] = "pcs"
    sheet6['G20'] = "TVA"
    sheet6['H20'] = 300000
    
    # Sauvegarder le fichier
    filename = "test_clients_validation.xlsx"
    workbook.save(filename)
    
    print(f"✅ Fichier de test créé: {filename}")
    print("\n📋 Scénarios de test:")
    print("  • Facture_Client_B2B_OK: Client B2B existant et actif")
    print("  • Facture_Client_Inexistant: Client non trouvé en base (erreur)")
    print("  • Facture_Client_Inactif: Client existant mais inactif (erreur)")
    print("  • Facture_Divers_Avec_NCC: Client divers avec NCC spécifique")
    print("  • Facture_Divers_Sans_NCC: Client divers sans NCC (template B2C par défaut)")
    print("  • Facture_B2B_Template_B2C: Client B2B avec template B2C")
    
    return filename

def main():
    print("🚀 Création des données de test pour validation clients...")
    
    # Créer la base de données de test
    db_path = create_clients_test_db()
    
    # Créer le fichier Excel de test
    excel_file = create_clients_validation_test()
    
    print(f"\n✅ Test prêt!")
    print(f"📄 Base de données: {db_path}")
    print(f"📄 Fichier Excel: {excel_file}")
    print(f"\n💡 Pour tester: python validate_sage100_structure.py {excel_file} {db_path}")

if __name__ == "__main__":
    main()
