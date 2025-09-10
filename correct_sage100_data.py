#!/usr/bin/env python3
"""
Correcteur automatique pour les données Sage 100
Corrige automatiquement les erreurs courantes dans les fichiers Excel
"""

import openpyxl
import re
import sys
from datetime import datetime

class Sage100DataCorrector:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.worksheet = self.workbook.active
        self.corrections = []
        
    def fix_ncc_format(self):
        """Corrige le format des NCC invalides"""
        print("🔧 Correction des formats NCC...")
        
        for row_num in range(2, self.worksheet.max_row + 1):
            ncc_value = self.worksheet[f'A{row_num}'].value
            
            if ncc_value and isinstance(ncc_value, str):
                # Nettoyer le NCC
                clean_ncc = re.sub(r'[^0-9A-Z]', '', ncc_value.upper())
                
                # Cas spéciaux
                if ncc_value.startswith('CC:'):
                    clean_ncc = ncc_value.replace('CC:', '').strip()
                    clean_ncc = re.sub(r'[^0-9A-Z]', '', clean_ncc.upper())
                elif ncc_value.startswith('N°'):
                    # Pour les formats comme "N°9015", générer un NCC valide
                    numbers = re.findall(r'\d+', ncc_value)
                    if numbers:
                        # Compléter à 7 chiffres et ajouter une lettre
                        number = numbers[0].zfill(7)
                        clean_ncc = number + 'A'
                
                # Vérifier si le format est maintenant valide
                if re.match(r'^\d{7}[A-Z]$', clean_ncc):
                    if clean_ncc != ncc_value:
                        self.worksheet[f'A{row_num}'].value = clean_ncc
                        self.corrections.append(f"Ligne {row_num}: NCC corrigé de '{ncc_value}' vers '{clean_ncc}'")
                elif len(clean_ncc) >= 7:
                    # Essayer de formater en 7 chiffres + lettre
                    if clean_ncc.isdigit():
                        formatted_ncc = clean_ncc[:7] + 'A'
                    else:
                        # Séparer chiffres et lettres
                        numbers = ''.join(filter(str.isdigit, clean_ncc))
                        letters = ''.join(filter(str.isalpha, clean_ncc))
                        
                        if len(numbers) >= 7 and letters:
                            formatted_ncc = numbers[:7] + letters[0]
                        elif len(numbers) >= 7:
                            formatted_ncc = numbers[:7] + 'A'
                        else:
                            formatted_ncc = numbers.zfill(7) + (letters[0] if letters else 'A')
                    
                    self.worksheet[f'A{row_num}'].value = formatted_ncc
                    self.corrections.append(f"Ligne {row_num}: NCC formaté de '{ncc_value}' vers '{formatted_ncc}'")
    
    def add_missing_payment_methods(self):
        """Ajoute des moyens de paiement par défaut"""
        print("🔧 Ajout des moyens de paiement manquants...")
        
        for row_num in range(2, self.worksheet.max_row + 1):
            payment_method = self.worksheet[f'R{row_num}'].value
            
            if not payment_method or payment_method == '':
                # Définir un moyen de paiement par défaut basé sur le type de client
                client_code = self.worksheet[f'C{row_num}'].value
                
                if client_code == '1999':  # Client divers
                    default_payment = 'ESPECES'
                else:
                    default_payment = 'VIREMENT'
                
                self.worksheet[f'R{row_num}'].value = default_payment
                self.corrections.append(f"Ligne {row_num}: Moyen de paiement ajouté: {default_payment}")
    
    def add_missing_ncc_for_clients(self):
        """Ajoute des NCC manquants pour les clients B2B uniquement"""
        print("🔧 Ajout des NCC manquants pour clients B2B...")
        
        ncc_counter = 1000000
        
        for row_num in range(2, self.worksheet.max_row + 1):
            client_code = self.worksheet[f'C{row_num}'].value
            ncc_value = self.worksheet[f'A{row_num}'].value
            
            if client_code != '1999':
                # Client B2B - NCC obligatoire en A6
                if not ncc_value or ncc_value == '':
                    generated_ncc = f"{ncc_counter:07d}B"
                    self.worksheet[f'A{row_num}'].value = generated_ncc
                    self.corrections.append(f"Ligne {row_num}: NCC généré pour client B2B: {generated_ncc}")
                    ncc_counter += 1
            # Pour les clients divers (1999), le NCC est optionnel (template B2C par défaut)
    
    def recalculate_amounts(self):
        """Recalcule les montants HT basés sur les quantités et prix unitaires"""
        print("🔧 Recalcul des montants HT...")
        
        for row_num in range(2, self.worksheet.max_row + 1):
            try:
                quantity = self.worksheet[f'L{row_num}'].value or 0
                unit_price = self.worksheet[f'M{row_num}'].value or 0
                current_amount = self.worksheet[f'N{row_num}'].value or 0
                
                if isinstance(quantity, (int, float)) and isinstance(unit_price, (int, float)):
                    calculated_amount = quantity * unit_price
                    
                    # Vérifier si le montant actuel est différent
                    if abs(calculated_amount - current_amount) > 0.01:  # Tolérance de 1 centime
                        self.worksheet[f'N{row_num}'].value = calculated_amount
                        self.corrections.append(f"Ligne {row_num}: Montant HT recalculé: {current_amount} → {calculated_amount}")
                        
            except (ValueError, TypeError):
                continue
    
    def save_corrected_file(self, output_path=None):
        """Sauvegarde le fichier corrigé"""
        if not output_path:
            # Créer un nom de fichier avec timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = self.file_path.replace('.xlsx', '')
            output_path = f"{base_name}_corrige_{timestamp}.xlsx"
        
        self.workbook.save(output_path)
        return output_path
    
    def generate_correction_report(self):
        """Génère un rapport des corrections effectuées"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = f"rapport_corrections_{timestamp}.txt"
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("RAPPORT DE CORRECTIONS SAGE 100\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"Fichier traité: {self.file_path}\n")
            f.write(f"Date: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write(f"Nombre de corrections: {len(self.corrections)}\n\n")
            
            if self.corrections:
                f.write("DÉTAIL DES CORRECTIONS:\n")
                f.write("-" * 30 + "\n")
                for correction in self.corrections:
                    f.write(f"• {correction}\n")
            else:
                f.write("Aucune correction nécessaire.\n")
        
        return report_path
    
    def run_all_corrections(self):
        """Execute toutes les corrections"""
        print("🚀 Démarrage des corrections automatiques...")
        
        self.fix_ncc_format()
        self.add_missing_ncc_for_clients()
        self.add_missing_payment_methods()
        self.recalculate_amounts()
        
        # Sauvegarder le fichier corrigé
        corrected_file = self.save_corrected_file()
        
        # Générer le rapport
        report_file = self.generate_correction_report()
        
        print(f"\n✅ Corrections terminées!")
        print(f"📄 Fichier corrigé: {corrected_file}")
        print(f"📊 Rapport: {report_file}")
        print(f"🔧 {len(self.corrections)} correction(s) effectuée(s)")
        
        return corrected_file, report_file

def main():
    if len(sys.argv) != 2:
        print("Usage: python correct_sage100_data.py <fichier_excel>")
        return
    
    file_path = sys.argv[1]
    
    try:
        corrector = Sage100DataCorrector(file_path)
        corrected_file, report_file = corrector.run_all_corrections()
        
        print(f"\n🎯 RÉSUMÉ:")
        print(f"   Fichier original: {file_path}")
        print(f"   Fichier corrigé: {corrected_file}")
        print(f"   Rapport détaillé: {report_file}")
        
    except FileNotFoundError:
        print(f"❌ Erreur: Fichier '{file_path}' non trouvé")
    except Exception as e:
        print(f"❌ Erreur lors de la correction: {str(e)}")

if __name__ == "__main__":
    main()
