#!/usr/bin/env python3
"""
Créateur de fichier de test pour validation des règles de nommage des clients
"""

import openpyxl
from datetime import datetime

def create_client_naming_test():
    """Crée un fichier de test avec les bonnes règles de nommage"""
    
    workbook = openpyxl.Workbook()
    
    # Supprimer la feuille par défaut
    workbook.remove(workbook.active)
    
    # Test 1: Client B2B avec nom en ligne 11
    sheet1 = workbook.create_sheet("Client_B2B_Nom_L11")
    
    sheet1['A3'] = "556101"  # Numéro facture
    sheet1['A5'] = "2000"    # Code client B2B (≠ 1999)
    sheet1['A6'] = "1234567A"  # NCC client B2B (obligatoire)
    sheet1['A8'] = "2025-09-10"  # Date
    sheet1['A10'] = "Gestoci"  # Point de vente
    sheet1['A11'] = "ENTREPRISE ABC SARL"  # Nom client B2B (ligne 11)
    sheet1['A18'] = "VIREMENT"  # Moyen de paiement
    
    # Produits
    sheet1['B20'] = "PROD001"
    sheet1['C20'] = "Service consulting"
    sheet1['D20'] = 1000000
    sheet1['E20'] = 1
    sheet1['F20'] = "service"
    sheet1['G20'] = "TVA"
    sheet1['H20'] = 1000000
    
    # Test 2: Client divers avec NCC (nom en ligne 13)
    sheet2 = workbook.create_sheet("Client_Divers_Avec_NCC")
    
    sheet2['A3'] = "556102"  # Numéro facture
    sheet2['A5'] = "1999"    # Code client divers
    sheet2['A6'] = "2354552Q"  # NCC générique pour divers
    sheet2['A8'] = "2025-09-10"  # Date
    sheet2['A10'] = "Gestoci"  # Point de vente
    sheet2['A11'] = "DIVERS CLIENTS"  # Nom générique (ligne 11)
    sheet2['A13'] = "MARIE KOUAME"  # Nom réel client divers (ligne 13)
    sheet2['A15'] = "7890123D"  # NCC spécifique client divers (optionnel)
    sheet2['A18'] = "ESPECES"  # Moyen de paiement
    
    # Produits
    sheet2['B20'] = "PROD002"
    sheet2['C20'] = "Carburant"
    sheet2['D20'] = 750
    sheet2['E20'] = 50
    sheet2['F20'] = "litres"
    sheet2['G20'] = "TVA"
    sheet2['H20'] = 37500
    
    # Test 3: Client divers SANS NCC (template B2C par défaut)
    sheet3 = workbook.create_sheet("Client_Divers_Sans_NCC")
    
    sheet3['A3'] = "556103"  # Numéro facture
    sheet3['A5'] = "1999"    # Code client divers
    sheet3['A6'] = "2354552Q"  # NCC générique pour divers
    sheet3['A8'] = "2025-09-10"  # Date
    sheet3['A10'] = "Gestoci"  # Point de vente
    sheet3['A11'] = "DIVERS CLIENTS"  # Nom générique (ligne 11)
    sheet3['A13'] = "JEAN KOUASSI"  # Nom réel client divers (ligne 13)
    # A15 vide intentionnellement - template B2C par défaut
    sheet3['A18'] = "ESPECES"  # Moyen de paiement
    
    # Produits
    sheet3['B20'] = "PROD003"
    sheet3['C20'] = "Essence"
    sheet3['D20'] = 780
    sheet3['E20'] = 40
    sheet3['F20'] = "litres"
    sheet3['G20'] = "TVA"
    sheet3['H20'] = 31200
    
    # Test 4: Client B2B sans nom en ligne 11 (erreur)
    sheet4 = workbook.create_sheet("Client_B2B_Sans_Nom")
    
    sheet4['A3'] = "556104"  # Numéro facture
    sheet4['A5'] = "3000"    # Code client B2B
    sheet4['A6'] = "9876543Z"  # NCC client B2B
    sheet4['A8'] = "2025-09-10"  # Date
    sheet4['A10'] = "Gestoci"  # Point de vente
    # A11 vide intentionnellement - ERREUR pour client B2B
    sheet4['A18'] = "VIREMENT"  # Moyen de paiement
    
    # Produits
    sheet4['B20'] = "PROD004"
    sheet4['C20'] = "Formation"
    sheet4['D20'] = 500000
    sheet4['E20'] = 2
    sheet4['F20'] = "session"
    sheet4['G20'] = "TVA"
    sheet4['H20'] = 1000000
    
    # Test 5: Client divers sans nom en ligne 13 (erreur)
    sheet5 = workbook.create_sheet("Client_Divers_Sans_Nom")
    
    sheet5['A3'] = "556105"  # Numéro facture
    sheet5['A5'] = "1999"    # Code client divers
    sheet5['A6'] = "2354552Q"  # NCC générique pour divers
    sheet5['A8'] = "2025-09-10"  # Date
    sheet5['A10'] = "Gestoci"  # Point de vente
    sheet5['A11'] = "DIVERS CLIENTS"  # Nom générique (ligne 11)
    # A13 vide intentionnellement - ERREUR pour client divers
    sheet5['A15'] = "5555555K"  # NCC spécifique
    sheet5['A18'] = "ESPECES"  # Moyen de paiement
    
    # Produits
    sheet5['B20'] = "PROD005"
    sheet5['C20'] = "Gasoil"
    sheet5['D20'] = 800
    sheet5['E20'] = 30
    sheet5['F20'] = "litres"
    sheet5['G20'] = "TVA"
    sheet5['H20'] = 24000
    
    # Sauvegarder le fichier
    filename = "test_client_naming.xlsx"
    workbook.save(filename)
    
    print(f"✅ Fichier de test créé: {filename}")
    print("\n📋 Contenu du test:")
    print("  • Client_B2B_Nom_L11: Client B2B avec nom correct en ligne 11")
    print("  • Client_Divers_Avec_NCC: Client divers avec NCC et nom en ligne 13")
    print("  • Client_Divers_Sans_NCC: Client divers SANS NCC (template B2C par défaut)")
    print("  • Client_B2B_Sans_Nom: Client B2B SANS nom en ligne 11 (erreur)")
    print("  • Client_Divers_Sans_Nom: Client divers SANS nom en ligne 13 (erreur)")
    
    return filename

if __name__ == "__main__":
    create_client_naming_test()
