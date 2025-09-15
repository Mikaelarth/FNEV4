#!/usr/bin/env python3
"""
Création d'un fichier Excel de test avec des doublons internes pour valider la détection
"""
import openpyxl
from datetime import datetime
import os

def create_test_file_with_duplicates():
    """Créer un fichier Excel avec des doublons internes pour tester"""
    
    # Créer un nouveau workbook
    wb = openpyxl.Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # Données de test avec doublons intentionnels
    factures_data = [
        {
            'feuille': 'Facture_DOUBLON_1',
            'numero': '556589',  # DOUBLON INTENTIONNEL
            'client_code': '3001A1',
            'client_nom': 'VENTES COMPT(S)',
            'date': '01/03/2025',
            'montant_ht': 852657.00,
            'montant_ttc': 1006135.26,
            'produits': [
                {'code': 'P001', 'designation': 'Produit Test 1', 'prix': 100.00, 'quantite': 1000},
                {'code': 'P002', 'designation': 'Produit Test 2', 'prix': 200.00, 'quantite': 2000},
            ]
        },
        {
            'feuille': 'Facture_DOUBLON_2',
            'numero': '556589',  # MÊME NUMÉRO = DOUBLON
            'client_code': '3001A1',
            'client_nom': 'VENTES COMPT(S)',
            'date': '01/03/2025',
            'montant_ht': 852657.00,
            'montant_ttc': 1006135.26,
            'produits': [
                {'code': 'P003', 'designation': 'Produit Test 3', 'prix': 150.00, 'quantite': 1500},
            ]
        },
        {
            'feuille': 'Facture_DOUBLON_3',
            'numero': '556589',  # ENCORE LE MÊME = TRIPLON !
            'client_code': '3001A1',
            'client_nom': 'VENTES COMPT(S)',
            'date': '01/03/2025',
            'montant_ht': 852657.00,
            'montant_ttc': 1006135.26,
            'produits': [
                {'code': 'P004', 'designation': 'Produit Test 4', 'prix': 300.00, 'quantite': 800},
            ]
        },
        {
            'feuille': 'Facture_UNIQUE',
            'numero': '999888',  # Facture unique
            'client_code': '1999',
            'client_nom': 'ARTHUR LE GRAND',
            'date': '11/08/2025',
            'montant_ht': 2332188.00,
            'montant_ttc': 2751981.84,
            'produits': [
                {'code': 'P005', 'designation': 'Produit Unique', 'prix': 500.00, 'quantite': 4000},
            ]
        },
        {
            'feuille': 'Facture_AUTRE_DOUBLON_A',
            'numero': '777555',  # Autre doublon
            'client_code': '1999',
            'client_nom': 'SIPARCOCI',
            'date': '21/03/2023',
            'montant_ht': 157390.00,
            'montant_ttc': 185720.20,
            'produits': [
                {'code': 'P006', 'designation': 'Autre Produit A', 'prix': 250.00, 'quantite': 600},
            ]
        },
        {
            'feuille': 'Facture_AUTRE_DOUBLON_B',
            'numero': '777555',  # Même numéro que précédent
            'client_code': '1999',
            'client_nom': 'SIPARCOCI',
            'date': '21/03/2023',
            'montant_ht': 157390.00,
            'montant_ttc': 185720.20,
            'produits': [
                {'code': 'P007', 'designation': 'Autre Produit B', 'prix': 180.00, 'quantite': 850},
            ]
        }
    ]
    
    for facture in factures_data:
        # Créer une nouvelle feuille
        ws = wb.create_sheet(title=facture['feuille'])
        
        # Structure selon exemple_structure_excel.py
        ws['A1'] = 'FACTURE SAGE 100'
        ws['A3'] = facture['numero']  # Numéro facture
        ws['A5'] = facture['client_code']  # Code client
        ws['A7'] = facture['client_nom']  # Nom client
        ws['A8'] = facture['date']  # Date facture
        ws['A10'] = 'DÉTAIL'
        
        # En-têtes produits (ligne 19)
        ws['A19'] = 'Ligne'
        ws['B19'] = 'Code'
        ws['C19'] = 'Désignation'
        ws['D19'] = 'Prix Unit.'
        ws['E19'] = 'Quantité'
        ws['F19'] = 'Emballage'
        ws['G19'] = 'TVA'
        ws['H19'] = 'Montant'
        
        # Produits (à partir de la ligne 20)
        row = 20
        for produit in facture['produits']:
            ws[f'A{row}'] = row - 19  # Numéro ligne
            ws[f'B{row}'] = produit['code']
            ws[f'C{row}'] = produit['designation']
            ws[f'D{row}'] = produit['prix']
            ws[f'E{row}'] = produit['quantite']
            ws[f'F{row}'] = 'UN'  # Emballage par défaut
            ws[f'G{row}'] = 20.0   # TVA 20%
            ws[f'H{row}'] = produit['prix'] * produit['quantite']
            row += 1
    
    # Sauvegarder le fichier
    filename = r'd:\PROJET\FNEV4\test_doublons_internes.xlsx'
    wb.save(filename)
    
    print(f"✅ Fichier créé : {filename}")
    print("\n📊 CONTENU DU FICHIER :")
    print("  - Facture 556589 : TRIPLE DOUBLON (3 feuilles)")
    print("  - Facture 777555 : DOUBLON (2 feuilles)")  
    print("  - Facture 999888 : UNIQUE (1 feuille)")
    print("\n🎯 RÉSULTAT ATTENDU DANS L'APERÇU :")
    print("  - 0 factures valides (toutes marquées comme doublons)")
    print("  - Messages d'erreur détaillant les doublons internes")
    print("  - Statut 'invalide' pour toutes les factures dupliquées")
    
    return filename

if __name__ == "__main__":
    create_test_file_with_duplicates()