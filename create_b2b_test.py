#!/usr/bin/env python3
"""
Créateur de fichier de test pour validation B2B vs Client Divers
"""

import openpyxl
from datetime import datetime

def create_b2b_vs_divers_test():
    """Crée un fichier de test avec clients B2B et clients divers"""
    
    workbook = openpyxl.Workbook()
    
    # Supprimer la feuille par défaut
    workbook.remove(workbook.active)
    
    # Test 1: Client B2B avec NCC valide
    sheet1 = workbook.create_sheet("Facture_B2B_Valide")
    
    # Données B2B correctes
    sheet1['A3'] = "556001"  # Numéro facture
    sheet1['A5'] = "2000"    # Code client B2B (≠ 1999)
    sheet1['A6'] = "1234567A"  # NCC client B2B (obligatoire)
    sheet1['A8'] = "2025-09-10"  # Date
    sheet1['A10'] = "Gestoci"  # Point de vente
    sheet1['A11'] = "ENTREPRISE ABC SARL"  # Nom client B2B
    sheet1['A18'] = "VIREMENT"  # Moyen de paiement
    
    # Produits
    sheet1['B20'] = "PROD001"
    sheet1['C20'] = "Service consulting"
    sheet1['D20'] = 1000000
    sheet1['E20'] = 1
    sheet1['F20'] = "service"
    sheet1['G20'] = "TVA"
    sheet1['H20'] = 1000000
    
    # Test 2: Client B2B SANS NCC (erreur)
    sheet2 = workbook.create_sheet("Facture_B2B_Sans_NCC")
    
    sheet2['A3'] = "556002"  # Numéro facture
    sheet2['A5'] = "3000"    # Code client B2B
    # A6 vide intentionnellement - ERREUR!
    sheet2['A8'] = "2025-09-10"  # Date
    sheet2['A10'] = "Gestoci"  # Point de vente
    sheet2['A11'] = "ENTREPRISE XYZ SARL"  # Nom client B2B
    sheet2['A18'] = "VIREMENT"  # Moyen de paiement
    
    # Produits
    sheet2['B20'] = "PROD002"
    sheet2['C20'] = "Formation"
    sheet2['D20'] = 500000
    sheet2['E20'] = 2
    sheet2['F20'] = "session"
    sheet2['G20'] = "TVA"
    sheet2['H20'] = 1000000
    
    # Test 3: Client Divers avec NCC valide
    sheet3 = workbook.create_sheet("Facture_Divers_Valide")
    
    sheet3['A3'] = "556003"  # Numéro facture
    sheet3['A5'] = "1999"    # Code client divers
    sheet3['A6'] = "2354552Q"  # NCC générique pour divers
    sheet3['A8'] = "2025-09-10"  # Date
    sheet3['A10'] = "Gestoci"  # Point de vente
    sheet3['A11'] = "DIVERS CLIENTS"  # Nom générique
    sheet3['A13'] = "MARIE KOUAME"  # Nom réel client divers
    sheet3['A15'] = "7890123D"  # NCC spécifique client divers
    sheet3['A18'] = "ESPECES"  # Moyen de paiement
    
    # Produits
    sheet3['B20'] = "PROD003"
    sheet3['C20'] = "Carburant"
    sheet3['D20'] = 750
    sheet3['E20'] = 50
    sheet3['F20'] = "litres"
    sheet3['G20'] = "TVA"
    sheet3['H20'] = 37500
    
    # Test 4: Client Divers SANS NCC spécifique (erreur)
    sheet4 = workbook.create_sheet("Facture_Divers_Sans_NCC")
    
    sheet4['A3'] = "556004"  # Numéro facture
    sheet4['A5'] = "1999"    # Code client divers
    sheet4['A6'] = "2354552Q"  # NCC générique pour divers
    sheet4['A8'] = "2025-09-10"  # Date
    sheet4['A10'] = "Gestoci"  # Point de vente
    sheet4['A11'] = "DIVERS CLIENTS"  # Nom générique
    sheet4['A13'] = "JEAN KOUASSI"  # Nom réel client divers
    # A15 vide intentionnellement - ERREUR!
    sheet4['A18'] = "ESPECES"  # Moyen de paiement
    
    # Produits
    sheet4['B20'] = "PROD004"
    sheet4['C20'] = "Essence"
    sheet4['D20'] = 780
    sheet4['E20'] = 40
    sheet4['F20'] = "litres"
    sheet4['G20'] = "TVA"
    sheet4['H20'] = 31200
    
    # Test 5: Client B2B avec NCC invalide (format incorrect)
    sheet5 = workbook.create_sheet("Facture_B2B_NCC_Invalide")
    
    sheet5['A3'] = "556005"  # Numéro facture
    sheet5['A5'] = "4000"    # Code client B2B
    sheet5['A6'] = "123ABC"  # NCC invalide - format incorrect
    sheet5['A8'] = "2025-09-10"  # Date
    sheet5['A10'] = "Gestoci"  # Point de vente
    sheet5['A11'] = "ENTREPRISE DEF SARL"  # Nom client B2B
    sheet5['A18'] = "CHEQUE"  # Moyen de paiement
    
    # Produits
    sheet5['B20'] = "PROD005"
    sheet5['C20'] = "Maintenance"
    sheet5['D20'] = 200000
    sheet5['E20'] = 3
    sheet5['F20'] = "heures"
    sheet5['G20'] = "TVA"
    sheet5['H20'] = 600000
    
    # Sauvegarder le fichier
    filename = "test_b2b_vs_divers.xlsx"
    workbook.save(filename)
    
    print(f"✅ Fichier de test créé: {filename}")
    print("\n📋 Contenu du test:")
    print("  • Facture_B2B_Valide: Client B2B avec NCC correct")
    print("  • Facture_B2B_Sans_NCC: Client B2B SANS NCC (erreur attendue)")
    print("  • Facture_Divers_Valide: Client divers avec NCC correct")
    print("  • Facture_Divers_Sans_NCC: Client divers SANS NCC spécifique (erreur attendue)")
    print("  • Facture_B2B_NCC_Invalide: Client B2B avec NCC format invalide (erreur attendue)")
    
    return filename

if __name__ == "__main__":
    create_b2b_vs_divers_test()
