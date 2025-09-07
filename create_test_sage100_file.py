#!/usr/bin/env python3
"""
Créer un fichier Excel de test simple pour vérifier le tooltip des articles
Structure Sage 100 v15 simplifiée pour debug
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def create_sage100_test_file():
    # Créer un nouveau classeur
    wb = Workbook()
    
    # Renommer la feuille par défaut au lieu de la supprimer
    ws = wb.active
    ws.title = "Facture test"
    
    # EN-TÊTE
    ws['A3'] = "556001"  # Numéro facture
    ws['A5'] = "CLIENT01"  # Code client
    ws['A6'] = "123456"  # NCC client
    ws['A8'] = "01/03/2025"  # Date facture
    ws['A10'] = "POINT1"  # Point de vente
    ws['A11'] = "CLIENT TEST SARL"  # Intitulé client
    ws['A18'] = "cash"  # Moyen de paiement A18
    
    # PRODUITS (à partir ligne 20)
    produits_test = [
        # Code, Désignation, Prix unitaire, Quantité, Emballage, TVA, Montant HT
        ("PROD001", "Ordinateur portable Dell", 450000, 1, "unité", "TVA", 450000),
        ("PROD002", "Souris sans fil", 15000, 2, "pcs", "TVA", 30000),
        ("PROD003", "Clavier mécanique", 80000, 1, "unité", "TVA", 80000),
    ]
    
    for i, (code, designation, prix, qte, emb, tva, montant) in enumerate(produits_test):
        row = 20 + i
        ws[f'B{row}'] = code
        ws[f'C{row}'] = designation
        ws[f'D{row}'] = prix
        ws[f'E{row}'] = qte
        ws[f'F{row}'] = emb
        ws[f'G{row}'] = tva
        ws[f'H{row}'] = montant
    
    # Mise en forme basique
    for row in [3, 5, 6, 8, 10, 11, 18]:
        ws[f'A{row}'].font = Font(bold=True)
    
    # Sauvegarder
    filename = "test_sage100_articles.xlsx"
    wb.save(filename)
    print(f"✅ Fichier créé : {filename}")
    print("\nContenu du fichier :")
    print("📋 Facture 556001 - CLIENT TEST SARL")
    print("📦 Articles :")
    for code, designation, prix, qte, emb, tva, montant in produits_test:
        print(f"  • {code}: {designation} - {qte} {emb} × {prix:,} = {montant:,} FCFA")
    
    return filename

if __name__ == "__main__":
    print("=== CRÉATION FICHIER TEST SAGE 100 V15 ===\n")
    filename = create_sage100_test_file()
    print(f"\n🎯 Utilisez ce fichier pour tester les tooltips articles dans FNEV4")
    print(f"📂 Fichier : {filename}")
