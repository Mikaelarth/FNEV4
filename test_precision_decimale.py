#!/usr/bin/env python3
"""
Script pour tester la précision des décimales dans l'import Sage 100
"""

import openpyxl
from datetime import datetime
import os

def creer_test_precision_decimale():
    """Crée un fichier Excel spécifiquement pour tester la précision décimale"""
    
    # Créer un nouveau classeur
    wb = openpyxl.Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # === FACTURE TEST PRÉCISION ===
    ws = wb.create_sheet("Test_Precision_Decimale")
    
    # Entête
    ws['A3'] = "PREC001"  # Numéro facture
    ws['A5'] = "CLI_PREC"  # Code client
    ws['A6'] = "1234567P"  # NCC client
    ws['A8'] = "45512"  # Date
    ws['A10'] = "Test Précision"  # Point de vente
    ws['A11'] = "CLIENT TEST PRÉCISION DÉCIMALE"  # Nom client
    ws['A18'] = "cash"  # Moyen de paiement
    
    # Produits avec prix très précis pour tester
    produits_test = [
        ("PROD1", "Produit 1.23456", 1.23456, 1, "pcs", 18, 1.23456),
        ("PROD2", "Produit 99.99", 99.99, 2, "pcs", 18, 199.98),
        ("PROD3", "Produit 123.456789", 123.456789, 1, "pcs", 9, 123.456789),
        ("PROD4", "Produit 0.01", 0.01, 1000, "pcs", 0, 10.00),
        ("PROD5", "Produit 999999.99", 999999.99, 1, "pcs", 18, 999999.99),
        ("PROD6", "Produit précision max", 12345.6789, 2.5, "kg", 9, 30864.19725),
        ("PROD7", "Produit avec TVA vide", 5678.9123, 3, "litres", None, 17036.7369),
    ]
    
    for i, (code, design, prix, qte, embp, tva, montant) in enumerate(produits_test, 20):
        ws[f'B{i}'] = code
        ws[f'C{i}'] = design
        ws[f'D{i}'] = prix  # Prix avec toute la précision
        ws[f'E{i}'] = qte
        ws[f'F{i}'] = embp
        if tva is not None:
            ws[f'G{i}'] = tva
        ws[f'H{i}'] = montant  # Montant avec toute la précision
    
    # Sauvegarder le fichier
    nom_fichier = "test_precision_decimale.xlsx"
    wb.save(nom_fichier)
    
    print(f"✅ Fichier créé : {nom_fichier}")
    print("\n📊 Produits de test créés :")
    for i, (code, design, prix, qte, embp, tva, montant) in enumerate(produits_test, 1):
        tva_str = f"{tva}%" if tva is not None else "vide"
        print(f"  {i}. {code}: {prix:.6f} € (TVA: {tva_str})")
    
    return nom_fichier

if __name__ == "__main__":
    print("🔧 Création du fichier de test pour la précision décimale...")
    creer_test_precision_decimale()
    print("\n✅ Test prêt ! Vous pouvez maintenant importer ce fichier dans FNEV4.")
