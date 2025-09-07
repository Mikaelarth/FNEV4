#!/usr/bin/env python3
"""
Test de compilation et vérification du système d'import exceptionnel
"""

import os
import sys
import subprocess

def test_compilation():
    """Test la compilation du projet"""
    print("🔨 Test de compilation du projet...")
    
    try:
        result = subprocess.run(['dotnet', 'build'], 
                              capture_output=True, text=True, cwd=r'c:\wamp64\www\FNEV4')
        
        if result.returncode == 0:
            print("✅ Compilation réussie !")
            return True
        else:
            print(f"❌ Erreurs de compilation:\n{result.stderr}")
            return False
            
    except Exception as e:
        print(f"❌ Erreur lors de la compilation: {e}")
        return False

def check_special_files():
    """Vérifie que tous les fichiers du système spécial sont présents"""
    print("📁 Vérification des fichiers du système spécial...")
    
    required_files = [
        r'src\FNEV4.Application\Special\ImportSpecialExcelUseCase.cs',
        r'src\FNEV4.Infrastructure\Special\SpecialExcelImportService.cs',
        r'ANALYSE-COMPATIBILITE-EXCEL-EXCEPTIONNEL.md',
        r'SPECIAL-IMPORT-README.md',
        r'clients.xlsx'
    ]
    
    missing_files = []
    for file_path in required_files:
        full_path = os.path.join(r'c:\wamp64\www\FNEV4', file_path)
        if not os.path.exists(full_path):
            missing_files.append(file_path)
        else:
            print(f"✅ {file_path}")
    
    if missing_files:
        print(f"❌ Fichiers manquants: {missing_files}")
        return False
    
    return True

def analyze_excel_file():
    """Analyse rapide du fichier Excel avec openpyxl"""
    print("📊 Analyse rapide du fichier Excel...")
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(r'c:\wamp64\www\FNEV4\clients.xlsx', data_only=True)
        ws = wb.active
        
        if ws is None:
            print("❌ Impossible d'accéder à la feuille Excel")
            return False
        
        # Compter les clients potentiels (lignes avec données en colonne A à partir de L16)
        client_count = 0
        row = 16  # Ligne L16
        max_row = ws.max_row if ws.max_row else 1000
        
        while row <= max_row:
            cell_value = ws[f'A{row}'].value
            if cell_value and str(cell_value).strip():
                client_count += 1
                if client_count <= 3:  # Afficher les 3 premiers
                    code_client = ws[f'A{row}'].value
                    ncc = ws[f'B{row}'].value
                    nom = ws[f'E{row}'].value
                    print(f"  Client {client_count}: CODE={code_client}, NCC={ncc}, NOM={nom}")
            
            row += 3  # Espacement de 2 lignes + 1 pour la ligne suivante
            
            if row > 1000:  # Sécurité pour éviter une boucle infinie
                break
        
        print(f"✅ {client_count} clients détectés dans le format spécial")
        return client_count > 0
        
    except ImportError:
        print("⚠️ openpyxl non disponible, analyse Excel ignorée")
        return True
    except Exception as e:
        print(f"❌ Erreur lors de l'analyse Excel: {e}")
        return False

def main():
    """Fonction principale de test"""
    print("🧪 Test du système d'import exceptionnel FNEV4")
    print("=" * 50)
    
    all_tests_passed = True
    
    # Test 1: Vérification des fichiers
    if not check_special_files():
        all_tests_passed = False
    
    print()
    
    # Test 2: Compilation
    if not test_compilation():
        all_tests_passed = False
    
    print()
    
    # Test 3: Analyse Excel
    if not analyze_excel_file():
        all_tests_passed = False
    
    print()
    print("=" * 50)
    
    if all_tests_passed:
        print("🎉 TOUS LES TESTS SONT PASSES !")
        print("✅ Le système d'import exceptionnel est prêt à être testé")
        print()
        print("📋 Prochaines étapes:")
        print("  1. Intégrer ImportSpecialExcelUseCase dans le DI")
        print("  2. Créer un bouton d'interface pour l'import exceptionnel")
        print("  3. Tester l'import avec le fichier clients.xlsx")
        print("  4. Supprimer le système après utilisation (voir SPECIAL-IMPORT-README.md)")
    else:
        print("❌ CERTAINS TESTS ONT ECHOUE")
        print("🔧 Corrigez les problèmes avant de continuer")
    
    return 0 if all_tests_passed else 1

if __name__ == "__main__":
    sys.exit(main())
