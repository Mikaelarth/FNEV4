#!/usr/bin/env python3
"""
Analyseur pour le fichier Excel exceptionnel clients.xlsx
Structure spéciale : colonnes fixes avec lignes espacées de 2 lignes
"""

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️ openpyxl non disponible. Installation recommandée: pip install openpyxl")

import os

def analyze_special_excel_file():
    """Analyse la structure du fichier clients.xlsx exceptionnel"""
    
    file_path = "clients.xlsx"
    
    print("=== ANALYSE DU FICHIER EXCEL EXCEPTIONNEL ===")
    print(f"Fichier: {file_path}")
    print()
    
    # Chargement avec openpyxl pour analyser la structure exacte
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"Nom de la feuille: {ws.title}")
    print(f"Dimensions: {ws.max_row} lignes x {ws.max_column} colonnes")
    print()
    
    # Mapping des colonnes selon les specs
    column_mapping = {
        'A': 'CODE CLIENT',
        'B': 'NCC',
        'E': 'NOM', 
        'G': 'EMAIL',
        'I': 'TELEPHONE',
        'K': 'MODE DE REGLEMENT',
        'M': 'TYPE DE FACTURATION',
        'O': 'DEVISE'
    }
    
    print("=== STRUCTURE DES COLONNES ===")
    for col, desc in column_mapping.items():
        print(f"Colonne {col}: {desc}")
    print()
    
    # Analyser les en-têtes dans les premières lignes
    print("=== ANALYSE DES EN-TÊTES (lignes 1-15) ===")
    for row in range(1, 16):
        values = []
        for col in ['A', 'B', 'E', 'G', 'I', 'K', 'M', 'O']:
            cell_value = ws[f"{col}{row}"].value
            if cell_value:
                values.append(f"{col}: {cell_value}")
        if values:
            print(f"Ligne {row}: {' | '.join(values)}")
    print()
    
    # Analyser la zone de données (à partir de L16)
    print("=== ANALYSE DES DONNÉES (L13, L16, L19, L22...) ===")
    test_rows = [13, 16, 19, 22, 25, 28]  # Premier test + premiers clients réels
    
    for row_num in test_rows:
        if row_num <= ws.max_row:
            print(f"--- Ligne {row_num} ---")
            client_data = {}
            has_data = False
            
            for col, desc in column_mapping.items():
                cell_value = ws[f"{col}{row_num}"].value
                if cell_value:
                    client_data[desc] = cell_value
                    has_data = True
                    
            if has_data:
                for desc, value in client_data.items():
                    print(f"  {desc}: {value}")
            else:
                print("  (ligne vide)")
            print()
    
    # Détecter automatiquement tous les clients
    print("=== DÉTECTION AUTOMATIQUE DES CLIENTS ===")
    clients_found = []
    
    # Commencer à partir de la ligne 16 (premier client réel)
    for row in range(16, ws.max_row + 1, 3):  # Incrément de 3 (client + 2 lignes vides)
        client_data = {}
        has_significant_data = False
        
        for col, desc in column_mapping.items():
            cell_value = ws[f"{col}{row}"].value
            if cell_value:
                client_data[desc] = str(cell_value).strip()
                
        # Vérifier si on a au moins un CODE CLIENT ou un NOM
        if client_data.get('CODE CLIENT') or client_data.get('NOM'):
            has_significant_data = True
            
        if has_significant_data:
            clients_found.append({
                'ligne': row,
                'data': client_data
            })
    
    print(f"Nombre de clients détectés: {len(clients_found)}")
    print()
    
    for i, client in enumerate(clients_found, 1):
        print(f"Client {i} (ligne {client['ligne']}):")
        for desc, value in client['data'].items():
            print(f"  {desc}: {value}")
        print()
    
    return clients_found

if __name__ == "__main__":
    try:
        clients = analyze_special_excel_file()
        print(f"✅ Analyse terminée. {len(clients)} clients trouvés.")
    except Exception as e:
        print(f"❌ Erreur lors de l'analyse: {e}")
        import traceback
        traceback.print_exc()
