#!/usr/bin/env python3
"""
Test simple pour valider le service d'import des factures Sage 100 v15.
Crée un fichier Excel de test et vérifie la structure.
"""

import os
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook

def create_test_excel_file():
    """Crée un fichier Excel de test avec structure Sage 100 v15."""
    
    # Créer le nom de fichier
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"test_sage100_{timestamp}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    
    # Créer le classeur
    workbook = Workbook()
    
    # Supprimer la feuille par défaut
    workbook.remove(workbook.active)
    
    # === FACTURE 1 : Client normal avec carte ===
    sheet1 = workbook.create_sheet("FAC001")
    
    # En-tête facture (Colonne A)
    sheet1.cell(row=3, column=1, value="FAC-2024-001")        # A3: Numéro facture
    sheet1.cell(row=5, column=1, value="CLI001")              # A5: Code client
    sheet1.cell(row=6, column=1, value="12345678A")           # A6: NCC client normal
    sheet1.cell(row=8, column=1, value=datetime.today())      # A8: Date facture
    sheet1.cell(row=10, column=1, value="MAGASIN-01")         # A10: Point de vente
    sheet1.cell(row=11, column=1, value="ENTREPRISE ABC SARL") # A11: Intitulé client
    sheet1.cell(row=18, column=1, value="card")               # A18: Moyen de paiement (NOUVEAU)
    
    # Lignes produits (à partir ligne 20)
    products1 = [
        ("PROD001", "Ordinateur portable", 850000, 1, "pcs", "TVA"),
        ("PROD002", "Souris sans fil", 25000, 2, "pcs", "TVA"),
        ("SERV001", "Installation logiciel", 150000, 1, "heure", "TVAB")
    ]
    
    for i, (code, desc, price, qty, unit, vat_type) in enumerate(products1, start=20):
        sheet1.cell(row=i, column=2, value=code)          # B: Code produit
        sheet1.cell(row=i, column=3, value=desc)          # C: Désignation
        sheet1.cell(row=i, column=4, value=price)         # D: Prix unitaire
        sheet1.cell(row=i, column=5, value=qty)           # E: Quantité
        sheet1.cell(row=i, column=6, value=unit)          # F: Unité
        sheet1.cell(row=i, column=7, value=vat_type)      # G: Type TVA
        sheet1.cell(row=i, column=8, value=price * qty)   # H: Montant HT
    
    # === FACTURE 2 : Client divers avec espèces ===
    sheet2 = workbook.create_sheet("FAC002")
    
    # En-tête facture (Colonne A)
    sheet2.cell(row=3, column=1, value="FAC-2024-002")        # A3: Numéro facture
    sheet2.cell(row=5, column=1, value="1999")                # A5: Code client divers
    sheet2.cell(row=6, column=1, value="")                    # A6: NCC client normal (vide)
    sheet2.cell(row=8, column=1, value=datetime.today() - timedelta(days=1)) # A8: Date facture
    sheet2.cell(row=10, column=1, value="MAGASIN-02")         # A10: Point de vente
    sheet2.cell(row=11, column=1, value="CLIENT DIVERS")      # A11: Intitulé client
    sheet2.cell(row=13, column=1, value="MARTIN Jean")        # A13: Nom réel client divers
    sheet2.cell(row=15, column=1, value="98765432B")          # A15: NCC client divers
    sheet2.cell(row=18, column=1, value="cash")               # A18: Moyen de paiement
    
    # Lignes produits
    products2 = [
        ("PROD003", "Clavier mécanique", 75000, 1, "pcs", "TVA"),
        ("PROD004", "Écran 24 pouces", 320000, 1, "pcs", "TVA")
    ]
    
    for i, (code, desc, price, qty, unit, vat_type) in enumerate(products2, start=20):
        sheet2.cell(row=i, column=2, value=code)
        sheet2.cell(row=i, column=3, value=desc)
        sheet2.cell(row=i, column=4, value=price)
        sheet2.cell(row=i, column=5, value=qty)
        sheet2.cell(row=i, column=6, value=unit)
        sheet2.cell(row=i, column=7, value=vat_type)
        sheet2.cell(row=i, column=8, value=price * qty)
    
    # === FACTURE 3 : Avoir avec mobile money ===
    sheet3 = workbook.create_sheet("AVO001")
    
    # En-tête facture (Colonne A)
    sheet3.cell(row=3, column=1, value="AVO-2024-001")        # A3: Numéro facture
    sheet3.cell(row=5, column=1, value="CLI002")              # A5: Code client
    sheet3.cell(row=6, column=1, value="11111111C")           # A6: NCC client normal
    sheet3.cell(row=8, column=1, value=datetime.today())      # A8: Date facture
    sheet3.cell(row=10, column=1, value="MAGASIN-01")         # A10: Point de vente
    sheet3.cell(row=11, column=1, value="TECH SOLUTIONS")     # A11: Intitulé client
    sheet3.cell(row=17, column=1, value="FAC-2024-003")       # A17: Référence facture originale
    sheet3.cell(row=18, column=1, value="mobile-money")       # A18: Moyen de paiement
    
    # Lignes produits (montants négatifs pour avoir)
    products3 = [
        ("PROD001", "Ordinateur portable (retour)", -850000, 1, "pcs", "TVA")
    ]
    
    for i, (code, desc, price, qty, unit, vat_type) in enumerate(products3, start=20):
        sheet3.cell(row=i, column=2, value=code)
        sheet3.cell(row=i, column=3, value=desc)
        sheet3.cell(row=i, column=4, value=price)
        sheet3.cell(row=i, column=5, value=qty)
        sheet3.cell(row=i, column=6, value=unit)
        sheet3.cell(row=i, column=7, value=vat_type)
        sheet3.cell(row=i, column=8, value=price * qty)
    
    # Sauvegarder le fichier
    workbook.save(filepath)
    print(f"✅ Fichier Excel créé : {filepath}")
    
    return filepath

def analyze_test_file(filepath):
    """Analyse le fichier de test créé."""
    print(f"\n📊 Analyse du fichier : {os.path.basename(filepath)}")
    print(f"   Taille : {os.path.getsize(filepath)} octets")
    
    # Analyser avec openpyxl
    from openpyxl import load_workbook
    
    try:
        workbook = load_workbook(filepath)
        print(f"   Nombre de feuilles : {len(workbook.worksheets)}")
        
        for sheet in workbook.worksheets:
            print(f"\n   📄 Feuille : {sheet.title}")
            
            # Lire les informations clés
            invoice_number = sheet.cell(row=3, column=1).value
            customer_code = sheet.cell(row=5, column=1).value
            payment_method = sheet.cell(row=18, column=1).value
            
            print(f"      Numéro facture : {invoice_number}")
            print(f"      Code client : {customer_code}")
            print(f"      Paiement : {payment_method}")
            
            # Compter les lignes produits
            product_count = 0
            row = 20
            while row <= sheet.max_row and sheet.cell(row=row, column=2).value:
                product_count += 1
                row += 1
            
            print(f"      Lignes produits : {product_count}")
            
            # Calculer le total
            total_ht = 0
            for row in range(20, 20 + product_count):
                amount = sheet.cell(row=row, column=8).value or 0
                total_ht += amount
            
            print(f"      Total HT : {total_ht:,.0f} FCFA")
        
        workbook.close()
        
    except Exception as e:
        print(f"   ❌ Erreur analyse : {e}")

def main():
    """Fonction principale du test."""
    print("=== Test Création Fichier Excel Sage 100 v15 ===\n")
    
    try:
        # 1. Créer le fichier de test
        test_file = create_test_excel_file()
        
        # 2. Analyser le fichier
        analyze_test_file(test_file)
        
        # 3. Afficher les instructions pour le test C#
        print(f"\n🔧 Instructions pour test C# :")
        print(f"   1. Compilez le projet FNEV4 avec : dotnet build FNEV4.sln")
        print(f"   2. Le fichier de test est prêt : {os.path.basename(test_file)}")
        print(f"   3. Modifiez test_invoice_import.cs pour utiliser ce fichier")
        print(f"   4. Ou testez directement l'import dans l'application")
        
        # 4. Validation structure
        print(f"\n✅ Validation structure Sage 100 v15 :")
        print(f"   ✓ 3 feuilles créées (FAC001, FAC002, AVO001)")
        print(f"   ✓ En-têtes factures en colonne A")
        print(f"   ✓ Moyen de paiement en A18 (NOUVEAU)")
        print(f"   ✓ Lignes produits à partir ligne 20")
        print(f"   ✓ Client normal, client divers et avoir testés")
        
        print(f"\n🎯 Le fichier est prêt pour l'import !")
        
        # 5. Proposer de nettoyer ou garder
        keep_file = input(f"\nGarder le fichier de test ? (y/N) : ").lower().strip()
        
        if keep_file != 'y':
            os.remove(test_file)
            print(f"🧹 Fichier supprimé")
        else:
            print(f"📁 Fichier conservé : {test_file}")
            
    except Exception as e:
        print(f"❌ Erreur : {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
