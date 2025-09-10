#!/usr/bin/env python3
"""
Script de validation des données Sage 100 selon exemple_structure_excel.py
Vérifie la conformité des factures importées avec les spécifications FNE
"""

import sys
import os
import openpyxl
import sqlite3
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional
import re

class Sage100FactureValidator:
    """Validateur de factures Sage 100 selon la structure définie"""
    
    MOYENS_PAIEMENT_VALIDES = {
        'cash', 'card', 'mobile-money', 'bank-transfer', 'check', 'credit',
        'ESPECES', 'CARTE', 'MOBILE-MONEY', 'VIREMENT', 'CHEQUE', 'CREDIT',
        'especes', 'carte', 'virement', 'cheque'
    }
    
    CODES_TVA_VALIDES = {
        'TVA', 'TVAB', 'TVAC', 'TVAD'
    }
    
    def __init__(self, fichier_excel: str, db_path: str = None):
        self.fichier_excel = fichier_excel
        self.db_path = db_path or self.find_database()
        self.workbook = None
        self.erreurs_globales = []
        self.factures_valides = []
        self.factures_invalides = []
        
    def find_database(self) -> str:
        """Trouve la base de données FNEV4"""
        possible_paths = [
            "fnev4.db",
            "../fnev4.db", 
            "../../fnev4.db",
            "src/FNEV4.Infrastructure/fnev4.db",
            "../src/FNEV4.Infrastructure/fnev4.db"
        ]
        
        for path in possible_paths:
            if os.path.exists(path):
                return path
        
        # Si aucune DB trouvée, on continue sans validation client
        return None
    
    def get_client_info(self, code_client: str) -> Optional[Dict]:
        """Récupère les informations du client depuis la base de données"""
        if not self.db_path or not os.path.exists(self.db_path):
            return None
            
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Requête pour récupérer les infos client
            cursor.execute("""
                SELECT Id, ClientCode, Name, ClientNcc, DefaultTemplate, IsActive 
                FROM Clients 
                WHERE ClientCode = ?
            """, (code_client,))
            
            result = cursor.fetchone()
            conn.close()
            
            if result:
                return {
                    'id': result[0],
                    'code_client': result[1],
                    'nom_commercial': result[2],
                    'ncc': result[3],
                    'template': result[4],
                    'est_actif': bool(result[5])
                }
            return None
            
        except Exception as e:
            print(f"⚠️  Erreur accès base de données: {e}")
            return None
        
    def ouvrir_fichier(self) -> bool:
        """Ouvre le fichier Excel"""
        try:
            self.workbook = openpyxl.load_workbook(self.fichier_excel, data_only=True)
            return True
        except Exception as e:
            self.erreurs_globales.append(f"Impossible d'ouvrir le fichier: {e}")
            return False
    
    def obtenir_valeur_cellule(self, worksheet, cellule: str) -> str:
        """Obtient la valeur d'une cellule de manière sécurisée"""
        try:
            valeur = worksheet[cellule].value
            if valeur is None:
                return ""
            return str(valeur).strip()
        except:
            return ""
    
    def valider_date(self, date_str: str) -> Tuple[bool, Optional[datetime]]:
        """Valide et convertit une date Excel"""
        if not date_str:
            return False, None
            
        try:
            # Essayer format numérique Excel (jours depuis 1900)
            if date_str.isdigit():
                jours = int(date_str)
                # Excel considère 1 janvier 1900 comme jour 1 (mais compte à tort 1900 comme bissextile)
                date_excel = datetime(1899, 12, 30) + timedelta(days=jours)
                return True, date_excel
            
            # Essayer format ISO
            date_obj = datetime.fromisoformat(date_str.replace('/', '-'))
            return True, date_obj
            
        except:
            return False, None
    
    def valider_facture(self, worksheet, nom_feuille: str) -> Dict:
        """Valide une facture complète"""
        facture = {
            'nom_feuille': nom_feuille,
            'erreurs': [],
            'avertissements': [],
            'est_valide': True,
            'donnees': {}
        }
        
        # === VALIDATION DE L'ENTÊTE ===
        
        # A3: Numéro de facture (obligatoire)
        numero_facture = self.obtenir_valeur_cellule(worksheet, 'A3')
        if not numero_facture:
            facture['erreurs'].append("A3: Numéro de facture manquant")
            facture['est_valide'] = False
        else:
            facture['donnees']['numero_facture'] = numero_facture
        
        # A5: Code client (obligatoire)
        code_client = self.obtenir_valeur_cellule(worksheet, 'A5')
        if not code_client:
            facture['erreurs'].append("A5: Code client manquant")
            facture['est_valide'] = False
        else:
            facture['donnees']['code_client'] = code_client
            
            # Validation spécifique selon type de client
            if code_client == "1999":
                # Client divers - template B2C par défaut
                nom_reel = self.obtenir_valeur_cellule(worksheet, 'A13')
                ncc_divers = self.obtenir_valeur_cellule(worksheet, 'A15')
                
                if not nom_reel:
                    facture['erreurs'].append("A13: Nom réel du client divers manquant (obligatoire pour code 1999)")
                    facture['est_valide'] = False
                
                # NCC optionnel pour les clients divers (template B2C par défaut)
                if ncc_divers and not self.valider_ncc(ncc_divers):
                    facture['erreurs'].append(f"A15: NCC '{ncc_divers}' invalide (format attendu: 7 chiffres + 1 lettre)")
                    facture['est_valide'] = False
                elif not ncc_divers:
                    facture['avertissements'].append("A15: NCC client divers non spécifié (template B2C appliqué par défaut)")
                
                facture['donnees']['nom_client'] = nom_reel
                facture['donnees']['ncc_client'] = ncc_divers if ncc_divers else ""
                facture['donnees']['type_client'] = 'divers'
                facture['donnees']['template'] = 'B2C'  # Template par défaut pour clients divers
                
            else:
                # Client B2B - vérifier existence dans la base de données
                client_info = self.get_client_info(code_client)
                
                if not client_info:
                    facture['erreurs'].append(f"A5: Client '{code_client}' non trouvé dans la base de données")
                    facture['est_valide'] = False
                elif not client_info['est_actif']:
                    facture['erreurs'].append(f"A5: Client '{code_client}' inactif")
                    facture['est_valide'] = False
                else:
                    # Client trouvé - vérifier NCC obligatoire
                    ncc_client = self.obtenir_valeur_cellule(worksheet, 'A6')
                    if not ncc_client:
                        facture['erreurs'].append("A6: NCC client B2B manquant (obligatoire pour clients B2B)")
                        facture['est_valide'] = False
                    elif not self.valider_ncc(ncc_client):
                        facture['erreurs'].append(f"A6: NCC '{ncc_client}' invalide (format attendu: 7 chiffres + 1 lettre)")
                        facture['est_valide'] = False
                    
                    # Utiliser le nom de la ligne 11 pour les clients B2B
                    nom_client = self.obtenir_valeur_cellule(worksheet, 'A11')
                    
                    facture['donnees']['nom_client'] = nom_client
                    facture['donnees']['ncc_client'] = ncc_client
                    facture['donnees']['type_client'] = 'b2b'
                    facture['donnees']['template'] = client_info['template']
                    facture['donnees']['client_db'] = client_info
        
        # A8: Date (obligatoire)
        date_str = self.obtenir_valeur_cellule(worksheet, 'A8')
        if not date_str:
            facture['erreurs'].append("A8: Date facture manquante")
            facture['est_valide'] = False
        else:
            est_valide, date_obj = self.valider_date(date_str)
            if not est_valide:
                facture['erreurs'].append(f"A8: Date '{date_str}' invalide")
                facture['est_valide'] = False
            else:
                facture['donnees']['date_facture'] = date_obj
        
        # A10: Point de vente (optionnel)
        point_vente = self.obtenir_valeur_cellule(worksheet, 'A10')
        facture['donnees']['point_vente'] = point_vente
        
        # A18: Moyen de paiement (optionnel mais recommandé)
        moyen_paiement = self.obtenir_valeur_cellule(worksheet, 'A18')
        if moyen_paiement:
            if moyen_paiement.lower() not in self.MOYENS_PAIEMENT_VALIDES:
                facture['erreurs'].append(f"A18: Moyen de paiement '{moyen_paiement}' invalide. Valides: {', '.join(self.MOYENS_PAIEMENT_VALIDES)}")
                facture['est_valide'] = False
            else:
                facture['donnees']['moyen_paiement'] = moyen_paiement
        else:
            facture['avertissements'].append("A18: Moyen de paiement non spécifié (utilisera celui du client par défaut)")
        
        # A17: Numéro facture avoir (optionnel)
        num_avoir = self.obtenir_valeur_cellule(worksheet, 'A17')
        if num_avoir:
            facture['donnees']['numero_avoir'] = num_avoir
        
        # === VALIDATION DES PRODUITS ===
        produits = []
        montant_total_ht = 0
        
        # Parcourir à partir de la ligne 20
        for row in range(20, worksheet.max_row + 1):
            code_produit = self.obtenir_valeur_cellule(worksheet, f'B{row}')
            
            if not code_produit:
                continue  # Ligne vide, continuer
            
            produit = {'ligne': row, 'erreurs': []}
            
            # B: Code produit (obligatoire si ligne présente)
            produit['code_produit'] = code_produit
            
            # C: Désignation (obligatoire)
            designation = self.obtenir_valeur_cellule(worksheet, f'C{row}')
            if not designation:
                produit['erreurs'].append(f"C{row}: Désignation manquante")
                facture['est_valide'] = False
            produit['designation'] = designation
            
            # D: Prix unitaire (obligatoire)
            prix_str = self.obtenir_valeur_cellule(worksheet, f'D{row}')
            try:
                prix_unitaire = float(prix_str.replace(',', '.')) if prix_str else 0
                if prix_unitaire <= 0:
                    produit['erreurs'].append(f"D{row}: Prix unitaire invalide ou nul")
                    facture['est_valide'] = False
                produit['prix_unitaire'] = prix_unitaire
            except:
                produit['erreurs'].append(f"D{row}: Prix unitaire '{prix_str}' non numérique")
                facture['est_valide'] = False
                produit['prix_unitaire'] = 0
            
            # E: Quantité (obligatoire)
            qte_str = self.obtenir_valeur_cellule(worksheet, f'E{row}')
            try:
                quantite = float(qte_str.replace(',', '.')) if qte_str else 0
                if quantite <= 0:
                    produit['erreurs'].append(f"E{row}: Quantité invalide ou nulle")
                    facture['est_valide'] = False
                produit['quantite'] = quantite
            except:
                produit['erreurs'].append(f"E{row}: Quantité '{qte_str}' non numérique")
                facture['est_valide'] = False
                produit['quantite'] = 0
            
            # F: Emballage (optionnel)
            produit['emballage'] = self.obtenir_valeur_cellule(worksheet, f'F{row}')
            
            # G: Code TVA (obligatoire)
            code_tva = self.obtenir_valeur_cellule(worksheet, f'G{row}')
            if not code_tva:
                produit['erreurs'].append(f"G{row}: Code TVA manquant")
                facture['est_valide'] = False
            elif code_tva not in self.CODES_TVA_VALIDES:
                produit['erreurs'].append(f"G{row}: Code TVA '{code_tva}' invalide. Valides: {', '.join(self.CODES_TVA_VALIDES)}")
                facture['est_valide'] = False
            produit['code_tva'] = code_tva
            
            # H: Montant HT (obligatoire)
            montant_str = self.obtenir_valeur_cellule(worksheet, f'H{row}')
            try:
                montant_ht = float(montant_str.replace(',', '.')) if montant_str else 0
                produit['montant_ht'] = montant_ht
                montant_total_ht += montant_ht
                
                # Vérification cohérence montant = prix × quantité
                montant_calcule = produit['prix_unitaire'] * produit['quantite']
                if abs(montant_ht - montant_calcule) > 0.01:  # Tolérance 1 centime
                    produit['erreurs'].append(f"H{row}: Montant HT incohérent (calculé: {montant_calcule:.2f}, fourni: {montant_ht:.2f})")
                    facture['avertissements'].append(f"Ligne {row}: Vérifier calcul montant HT")
                
            except:
                produit['erreurs'].append(f"H{row}: Montant HT '{montant_str}' non numérique")
                facture['est_valide'] = False
                produit['montant_ht'] = 0
            
            produits.append(produit)
        
        facture['donnees']['produits'] = produits
        facture['donnees']['nombre_produits'] = len(produits)
        facture['donnees']['montant_total_ht'] = montant_total_ht
        facture['donnees']['montant_total_ttc'] = montant_total_ht * 1.18  # Estimation 18% TVA
        
        # Validation globale
        if len(produits) == 0:
            facture['erreurs'].append("Aucun produit trouvé (données manquantes à partir de la ligne 20)")
            facture['est_valide'] = False
        
        return facture
    
    def valider_ncc(self, ncc: str) -> bool:
        """Valide un NCC (Numéro de Compte Contribuable)"""
        if not ncc:
            return False
        
        # Format attendu: 7 chiffres + 1 lettre
        pattern = r'^\d{7}[A-Z]$'
        return bool(re.match(pattern, ncc.upper()))
    
    def valider_fichier_complet(self) -> Dict:
        """Valide le fichier Excel complet"""
        if not self.ouvrir_fichier():
            return {
                'succes': False,
                'erreurs_globales': self.erreurs_globales,
                'factures_valides': [],
                'factures_invalides': []
            }
        
        print(f"📁 Validation du fichier: {self.fichier_excel}")
        print(f"📊 Nombre de feuilles: {len(self.workbook.sheetnames)}")
        print()
        
        for nom_feuille in self.workbook.sheetnames:
            print(f"🔍 Validation feuille: {nom_feuille}")
            worksheet = self.workbook[nom_feuille]
            
            facture = self.valider_facture(worksheet, nom_feuille)
            
            if facture['est_valide']:
                self.factures_valides.append(facture)
                print(f"   ✅ Valide - {facture['donnees'].get('nombre_produits', 0)} produits")
            else:
                self.factures_invalides.append(facture)
                print(f"   ❌ Invalide - {len(facture['erreurs'])} erreurs")
                for erreur in facture['erreurs'][:3]:  # Afficher max 3 erreurs
                    print(f"      • {erreur}")
                if len(facture['erreurs']) > 3:
                    print(f"      • ... et {len(facture['erreurs']) - 3} autres erreurs")
            
            if facture['avertissements']:
                print(f"   ⚠️  {len(facture['avertissements'])} avertissement(s)")
            
            print()
        
        return {
            'succes': True,
            'nombre_feuilles': len(self.workbook.sheetnames),
            'factures_valides': self.factures_valides,
            'factures_invalides': self.factures_invalides,
            'erreurs_globales': self.erreurs_globales
        }
    
    def generer_rapport(self, resultat: Dict) -> str:
        """Génère un rapport détaillé de validation"""
        rapport = []
        rapport.append("=" * 80)
        rapport.append("RAPPORT DE VALIDATION SAGE 100")
        rapport.append("=" * 80)
        rapport.append(f"Fichier analysé: {self.fichier_excel}")
        rapport.append(f"Date validation: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        rapport.append("")
        
        if not resultat['succes']:
            rapport.append("❌ ÉCHEC DE LA VALIDATION")
            for erreur in resultat['erreurs_globales']:
                rapport.append(f"   • {erreur}")
            return "\n".join(rapport)
        
        # Statistiques générales
        total_factures = len(resultat['factures_valides']) + len(resultat['factures_invalides'])
        rapport.append("📊 STATISTIQUES GÉNÉRALES")
        rapport.append("-" * 40)
        rapport.append(f"Nombre total de feuilles: {resultat['nombre_feuilles']}")
        rapport.append(f"Factures valides: {len(resultat['factures_valides'])}")
        rapport.append(f"Factures invalides: {len(resultat['factures_invalides'])}")
        rapport.append(f"Taux de réussite: {len(resultat['factures_valides'])/total_factures*100:.1f}%")
        rapport.append("")
        
        # Détail des factures valides
        if resultat['factures_valides']:
            rapport.append("✅ FACTURES VALIDES")
            rapport.append("-" * 40)
            total_ht = 0
            total_produits = 0
            
            for facture in resultat['factures_valides']:
                donnees = facture['donnees']
                rapport.append(f"• {facture['nom_feuille']}")
                rapport.append(f"  N°: {donnees.get('numero_facture', 'N/A')}")
                rapport.append(f"  Client: {donnees.get('nom_client', 'N/A')} ({donnees.get('code_client', 'N/A')})")
                rapport.append(f"  Template: {donnees.get('template', 'N/A')}")
                rapport.append(f"  Produits: {donnees.get('nombre_produits', 0)}")
                rapport.append(f"  Montant HT: {donnees.get('montant_total_ht', 0):,.2f} FCFA")
                
                total_ht += donnees.get('montant_total_ht', 0)
                total_produits += donnees.get('nombre_produits', 0)
                
                if facture['avertissements']:
                    rapport.append("  ⚠️  Avertissements:")
                    for avert in facture['avertissements']:
                        rapport.append(f"     - {avert}")
                rapport.append("")
            
            rapport.append(f"TOTAL GÉNÉRAL: {total_ht:,.2f} FCFA HT ({total_produits} produits)")
            rapport.append("")
        
        # Détail des factures invalides
        if resultat['factures_invalides']:
            rapport.append("❌ FACTURES INVALIDES")
            rapport.append("-" * 40)
            
            for facture in resultat['factures_invalides']:
                rapport.append(f"• {facture['nom_feuille']}")
                rapport.append("  Erreurs:")
                for erreur in facture['erreurs']:
                    rapport.append(f"     - {erreur}")
                
                if facture['avertissements']:
                    rapport.append("  Avertissements:")
                    for avert in facture['avertissements']:
                        rapport.append(f"     - {avert}")
                rapport.append("")
        
        return "\n".join(rapport)


def main():
    if len(sys.argv) < 2:
        print("Usage: python validate_sage100_structure.py <fichier_excel> [db_path]")
        print("\nExemple: python validate_sage100_structure.py factures.xlsx")
        print("         python validate_sage100_structure.py factures.xlsx test_clients.db")
        sys.exit(1)
    
    fichier_excel = sys.argv[1]
    db_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    if not os.path.exists(fichier_excel):
        print(f"❌ Fichier non trouvé: {fichier_excel}")
        sys.exit(1)
    
    print("🚀 Démarrage de la validation Sage 100...")
    print()
    
    validator = Sage100FactureValidator(fichier_excel, db_path)
    resultat = validator.valider_fichier_complet()
    
    # Affichage du rapport
    rapport = validator.generer_rapport(resultat)
    print(rapport)
    
    # Sauvegarde du rapport
    nom_rapport = f"rapport_validation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    with open(nom_rapport, 'w', encoding='utf-8') as f:
        f.write(rapport)
    
    print(f"📄 Rapport sauvegardé: {nom_rapport}")
    
    # Code de sortie
    if resultat['succes'] and len(resultat['factures_invalides']) == 0:
        print("\n🎉 VALIDATION RÉUSSIE - Toutes les factures sont conformes!")
        sys.exit(0)
    else:
        print(f"\n⚠️  VALIDATION PARTIELLE - {len(resultat.get('factures_invalides', []))} facture(s) à corriger")
        sys.exit(1)


if __name__ == "__main__":
    main()
