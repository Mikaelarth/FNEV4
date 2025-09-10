#!/usr/bin/env python3
"""
Script de création d'un fichier Excel de test conforme à exemple_structure_excel.py
Pour valider le fonctionnement du validateur Sage 100
"""

import openpyxl
from datetime import datetime, timedelta
import os

def creer_fichier_test_sage100():
    """Crée un fichier Excel de test avec plusieurs factures"""
    
    # Créer un nouveau classeur
    wb = openpyxl.Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # === FACTURE 1: Client normal valide ===
    ws1 = wb.create_sheet("Facture_001")
    
    # Entête
    ws1['A3'] = "FAC001"  # Numéro facture
    ws1['A5'] = "CLI001"  # Code client normal
    ws1['A6'] = "1234567A"  # NCC client
    ws1['A8'] = "45508"  # Date en format Excel (01/08/2024)
    ws1['A10'] = "Abidjan"  # Point de vente
    ws1['A11'] = "ENTREPRISE TECH SARL"  # Nom client
    ws1['A18'] = "cash"  # Moyen de paiement
    
    # Produits
    ws1['B20'] = "ORD001"
    ws1['C20'] = "Ordinateur portable Dell"
    ws1['D20'] = 849999.50
    ws1['E20'] = 1
    ws1['F20'] = "pcs"
    ws1['G20'] = 18  # TVA 18%
    ws1['H20'] = 849999.50
    
    ws1['B21'] = "SOU001"
    ws1['C21'] = "Souris sans fil"
    ws1['D21'] = 24750.25
    ws1['E21'] = 2
    ws1['F21'] = "pcs"
    ws1['G21'] = 18  # TVA 18%
    ws1['H21'] = 49500.50
    
    # === FACTURE 2: Client divers (1999) valide ===
    ws2 = wb.create_sheet("Facture_002")
    
    # Entête
    ws2['A3'] = "FAC002"  # Numéro facture
    ws2['A5'] = "1999"  # Code client divers
    ws2['A6'] = "2354552Q"  # NCC générique client divers
    ws2['A8'] = "45509"  # Date en format Excel (02/08/2024)
    ws2['A10'] = "Gestoci"  # Point de vente
    ws2['A11'] = "DIVERS CLIENTS CARBURANTS"  # Intitulé générique
    ws2['A13'] = "KOUAME JEAN-BAPTISTE"  # Nom réel client divers
    ws2['A15'] = "9876543Z"  # NCC spécifique client divers
    ws2['A18'] = "mobile-money"  # Moyen de paiement
    
    # Produits
    ws2['B20'] = "CARB001"
    ws2['C20'] = "Essence Super 95"
    ws2['D20'] = 649.75
    ws2['E20'] = 50
    ws2['F20'] = "litres"
    ws2['G20'] = 9  # TVAB 9%
    ws2['H20'] = 32487.50
    
    # === FACTURE 3: Facture avec erreurs pour tester la validation ===
    ws3 = wb.create_sheet("Facture_003_ERREURS")
    
    # Entête avec erreurs
    ws3['A3'] = ""  # Numéro facture manquant
    ws3['A5'] = "1999"  # Code client divers
    ws3['A6'] = "2354552Q"  # NCC générique
    ws3['A8'] = "date_invalide"  # Date invalide
    ws3['A10'] = "Bouaké"  # Point de vente
    ws3['A11'] = "DIVERS CLIENTS"  # Intitulé générique
    # A13 manquant (nom réel client divers)
    ws3['A15'] = "123ABC"  # NCC invalide (mauvais format)
    ws3['A18'] = "bitcoin"  # Moyen de paiement invalide
    
    # Produits avec erreurs
    ws3['B20'] = "PROD001"
    ws3['C20'] = ""  # Désignation manquante
    ws3['D20'] = "prix_invalide"  # Prix non numérique
    ws3['E20'] = 0  # Quantité nulle
    ws3['F20'] = "pcs"
    ws3['G20'] = 99  # Code TVA invalide (pourcentage inexistant)
    ws3['H20'] = 100000.00
    
    # === FACTURE 4: Facture avoir ===
    ws4 = wb.create_sheet("Avoir_001")
    
    # Entête
    ws4['A3'] = "AVO001"  # Numéro avoir
    ws4['A5'] = "CLI002"  # Code client
    ws4['A6'] = "7654321B"  # NCC client
    ws4['A8'] = "45510"  # Date
    ws4['A10'] = "Yamoussoukro"  # Point de vente
    ws4['A11'] = "SOCIÉTÉ MODERNE CI"  # Nom client
    ws4['A17'] = "FAC150"  # Numéro facture d'origine
    ws4['A18'] = "bank-transfer"  # Moyen de paiement
    
    # Produits (montants négatifs pour avoir)
    ws4['B20'] = "TEL001"
    ws4['C20'] = "Téléphone défectueux retourné"
    ws4['D20'] = -149999.99
    ws4['E20'] = 1
    ws4['F20'] = "pcs"
    ws4['G20'] = 18  # TVA 18%
    ws4['H20'] = -149999.99
    
    # === FACTURE 5: Facture complexe avec plusieurs produits ===
    ws5 = wb.create_sheet("Facture_COMPLEXE")
    
    # Entête
    ws5['A3'] = "FAC999"
    ws5['A5'] = "ENTREP001"
    ws5['A6'] = "1111111X"
    ws5['A8'] = "45511"
    ws5['A10'] = "GSM (Fictif)"
    ws5['A11'] = "GRANDE ENTREPRISE SARL"
    ws5['A18'] = "check"
    
    # Produits multiples
    produits = [
        ("SERV001", "Consultation technique", 99999.75, 1, "service", 18, 99999.75),
        ("MAT001", "Matériel informatique", 499850.50, 3, "pcs", 18, 1499551.50),
        ("FORM001", "Formation utilisateurs", 74999.25, 8, "heures", 9, 599994.00),
        ("MAINT001", "Maintenance préventive", 199950.75, 1, "service", 18, 199950.75),
        ("LIC001", "Licence logiciel", 349999.00, 2, "licence", None, 699998.00)  # Cellule TVA vide
    ]
    
    for i, (code, design, prix, qte, embp, tva, montant) in enumerate(produits, 20):
        ws5[f'B{i}'] = code
        ws5[f'C{i}'] = design
        ws5[f'D{i}'] = prix
        ws5[f'E{i}'] = qte
        ws5[f'F{i}'] = embp
        if tva is not None:  # Permettre les cellules vides pour TVA
            ws5[f'G{i}'] = tva
        ws5[f'H{i}'] = montant
    
    # === FACTURE 6: Test des cellules TVA vides et valeurs spécifiques ===
    ws6 = wb.create_sheet("Test_TVA_Cases")
    
    # Entête
    ws6['A3'] = "TEST001"
    ws6['A5'] = "CLI_TEST"
    ws6['A6'] = "1234567T"
    ws6['A8'] = "45512"
    ws6['A10'] = "Test"
    ws6['A11'] = "CLIENT TEST TVA"
    ws6['A18'] = "cash"
    
    # Produits pour tester différents cas TVA
    ws6['B20'] = "PROD_18"
    ws6['C20'] = "Produit avec TVA 18%"
    ws6['D20'] = 100.50
    ws6['E20'] = 1
    ws6['F20'] = "pcs"
    ws6['G20'] = 18  # TVA explicite 18%
    ws6['H20'] = 100.50
    
    ws6['B21'] = "PROD_9"
    ws6['C21'] = "Produit avec TVA 9%"
    ws6['D21'] = 200.75
    ws6['E21'] = 1
    ws6['F21'] = "pcs"
    ws6['G21'] = 9  # TVAB 9%
    ws6['H21'] = 200.75
    
    ws6['B22'] = "PROD_0"
    ws6['C22'] = "Produit avec TVA 0%"
    ws6['D22'] = 150.00
    ws6['E22'] = 1
    ws6['F22'] = "pcs"
    ws6['G22'] = 0  # TVAC 0%
    ws6['H22'] = 150.00
    
    ws6['B23'] = "PROD_VIDE"
    ws6['C23'] = "Produit avec cellule TVA vide"
    ws6['D23'] = 300.25
    ws6['E23'] = 1
    ws6['F23'] = "pcs"
    # G23 intentionnellement laissé vide
    ws6['H23'] = 300.25
    
    # Sauvegarder le fichier
    nom_fichier = "test_factures_sage100.xlsx"
    wb.save(nom_fichier)
    
    print(f"✅ Fichier de test créé: {nom_fichier}")
    print()
    print("📊 Contenu du fichier:")
    print("   • Facture_001: Client normal valide")
    print("   • Facture_002: Client divers (1999) valide")
    print("   • Facture_003_ERREURS: Facture avec erreurs de validation")
    print("   • Avoir_001: Facture avoir")
    print("   • Facture_COMPLEXE: Facture avec 5 produits différents")
    print()
    print("🔍 Pour valider ce fichier, exécutez:")
    print(f"   python validate_sage100_structure.py {nom_fichier}")
    
    return nom_fichier

if __name__ == "__main__":
    creer_fichier_test_sage100()
